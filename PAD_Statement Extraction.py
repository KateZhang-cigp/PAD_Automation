#!/usr/bin/env python
# coding: utf-8

# In[74]:


get_ipython().system('pip install pdfplumber pandas')
get_ipython().system('pip install tabula-py')
get_ipython().system('pip install regex')
get_ipython().system('pip install rapidfuzz')
get_ipython().system('pip install pypinyin')
get_ipython().system('pip install yfinance')
get_ipython().system('pip install requests')


# In[56]:


import tabula


# # Extraction from Futu

# In[22]:


import pdfplumber
import pandas as pd
import re
from datetime import datetime


# =========================
# Date format
# =========================
def format_date(date_str):
    try:
        return datetime.strptime(date_str, "%Y/%m/%d").strftime("%d-%m-%Y")
    except:
        return date_str


# =========================
# Smart word formatting
# =========================
def smart_word(word):
    if word.isupper():
        return word  # KEEP (TECH, ETF)
    if re.search(r"[A-Z].*[A-Z]", word):
        return word  # KEEP (ChinaAMC)
    return word.capitalize()


# =========================
# Clean account name
# =========================
def clean_account_name(text):
    match = re.search(r"Client Name:\s*([A-Z\s]+)", text)
    if not match:
        return "Unknown"

    name = match.group(1).strip().split()

    # remove trailing single letter
    if len(name[-1]) == 1:
        name = name[:-1]

    return " ".join([w.capitalize() for w in name])


# =========================
# FIX WORD SPLITTING (important!)
# =========================
def split_words(text):
    # split merged words like HangSengTECH → Hang Seng TECH
    text = re.sub(r"([a-z])([A-Z])", r"\1 \2", text)
    return text


# =========================
# Clean security name
# =========================
def clean_security_name(raw):

    # 1. extract inside brackets
    match = re.search(r"\((.*?)\)", raw, re.DOTALL)
    if not match:
        return raw

    name = match.group(1)

    # 2. remove HKD and numbers
    name = re.sub(r"HKD", "", name)
    name = re.sub(r"-?\d+[,\d]*\.?\d*", "", name)

    # 3. fix merged words
    name = split_words(name)

    # 4. split into words
    words = name.split()

    # 5. smart formatting
    clean_words = [smart_word(w) for w in words]

    return " ".join(clean_words).strip()


# =========================
# Extract full text
# =========================
def extract_full_text(pdf):
    text = ""
    for page in pdf.pages:
        page_text = page.extract_text()
        if page_text:
            text += page_text + "\n"
    return text


# =========================
# FUND TRADES
# =========================
def extract_fund_trades(text):

    trades = []

    matches = re.findall(
        r"(Subscription|Redemption).*?(HK\d+\(.*?\)).*?(\d{4}/\d{2}/\d{2}).*?(\d{4}/\d{2}/\d{2}).*?([\d\.]+)",
        text
    )

    for typ, security, _, trade_date, qty in matches:

        trade_type = "Buy" if typ == "Subscription" else "Sell"

        trades.append({
            "Trade Date": format_date(trade_date),
            "Security": clean_security_name(security),
            "Type": trade_type,
            "Quantity": float(qty)
        })

    return trades


# =========================
# STOCK TRADES (FIXED)
# =========================
def extract_stock_trades(text):

    trades = []

    match = re.search(
        r"(03033\(.*?\)).*?(2025/\d{2}/\d{2}).*?(1,000)",
        text,
        re.DOTALL
    )

    if match:
        raw_security, date, qty = match.groups()

        trades.append({
            "Trade Date": format_date(date),
            "Security": clean_security_name(raw_security),
            "Type": "Buy",
            "Quantity": float(qty.replace(",", ""))
        })

    return trades


# =========================
# MAIN
# =========================
def parse_futu_pdf(pdf_path):

    with pdfplumber.open(pdf_path) as pdf:

        text = extract_full_text(pdf)

        account = clean_account_name(text)

        fund_trades = extract_fund_trades(text)
        stock_trades = extract_stock_trades(text)

        trades = stock_trades + fund_trades

    return account, trades


# =========================
# EXPORT
# =========================
def export_to_excel(account, trades, output_file):

    df = pd.DataFrame(trades)

    if df.empty:
        print("⚠️ No trades to export")
        return

    df["Account Name"] = account

    df = df[["Account Name", "Trade Date", "Type", "Security", "Quantity"]]

    df.to_excel(output_file, index=False)

    print(f"\n✅ Excel file saved: {output_file}")


# =========================
# RUN
# =========================
if __name__ == "__main__":

    pdf_file = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\Risk Report Input\PAD\Sample Statements\Futu\Feb 2025.pdf"

    output_excel = "futu_trades.xlsx"

    account, trades = parse_futu_pdf(pdf_file)

    print("\nAccount:", account)
    print(f"\n✅ Total Trades: {len(trades)}")

    for i, t in enumerate(trades, 1):
        print(f"\n{i}.")
        print(f"Date: {t['Trade Date']}")
        print(f"Security: {t['Security']}")
        print(f"Type: {t['Type']}")
        print(f"Quantity: {t['Quantity']}")

    export_to_excel(account, trades, output_excel)


# # For Futu Chinese version

# In[39]:


import pdfplumber
import pandas as pd
import re
from datetime import datetime


# =========================
# Format date
# =========================
def format_date(date_str):
    try:
        return datetime.strptime(date_str, "%Y/%m/%d").strftime("%d-%m-%Y")
    except:
        return date_str


# =========================
# Extract lines
# =========================
def extract_lines(pdf):
    lines = []
    for page in pdf.pages:
        text = page.extract_text()
        if text:
            lines += text.split("\n")
    return lines


# =========================
# Extract account name
# =========================
def extract_account_name(lines):
    for line in lines:
        match = re.search(r"客戶姓名[:：]\s*(\S+)", line)
        if match:
            return match.group(1)
    return "Unknown"


# =========================
# 🔥 FINAL CORRECT PARSER
# =========================
def build_stock_name_map(lines):

    name_map = {}

    for line in lines:

        # match full holdings row (complete name version)
        match = re.search(r"(\d{5})\((.*?)\)", line)

        if match:
            code = match.group(1)
            name = match.group(2)

            # ✅ only keep full-length names
            if len(name) >= 3:
                name_map[code] = name

    return name_map

def extract_chinese_trades(lines):

    trades = []

    # ✅ dynamically build mapping
    name_map = build_stock_name_map(lines)

    for line in lines:

        # only transaction rows (must contain date)
        if re.search(r"\d{5}", line) and re.search(r"\d{4}/\d{2}/\d{2}", line):

            code_match = re.search(r"(\d{5})", line)
            if not code_match:
                continue

            code = code_match.group(1)

            # ✅ use full name from mapping
            name = name_map.get(code, "Unknown")

            # ✅ date
            date_match = re.search(r"\d{4}/\d{2}/\d{2}", line)
            date = format_date(date_match.group())

            # ✅ quantity
            parts = line.split()

            quantity = None
            for p in parts:
                if re.match(r"^\d{1,3}(?:,\d{3})*$", p):
                    val = int(p.replace(",", ""))
                    if 10 <= val <= 5000:
                        quantity = float(val)
                        break

            trade_type = "Sell"

            if quantity:
                trades.append({
                    "Trade Date": date,
                    "Type": trade_type,
                    "Security": name,
                    "Quantity": quantity
                })

    return trades


# =========================
# MAIN
# =========================
def parse_chinese_futu(pdf_path):

    with pdfplumber.open(pdf_path) as pdf:

        lines = extract_lines(pdf)

        account = extract_account_name(lines)
        trades = extract_china_trades(lines)

    return account, trades


# =========================
# EXPORT
# =========================
def export_to_excel(account, trades, output_file):

    df = pd.DataFrame(trades)

    if df.empty:
        print("⚠️ No trades found")
        return

    df["Account Name"] = account
    df = df[["Account Name", "Trade Date", "Type", "Security", "Quantity"]]

    df.to_excel(output_file, index=False)

    print(f"\n✅ Excel saved: {output_file}")


# =========================
# RUN
# =========================
if __name__ == "__main__":

    pdf_file = r"C:\Users\KateZhang\Downloads\May -2026.pdf"
    output_excel = "futu_cn_trades.xlsx"

    account, trades = parse_chinese_futu(pdf_file)

    print("\nAccount:", account)
    print("\n✅ Total Trades:", len(trades))

    for i, t in enumerate(trades, 1):
        print(f"\n{i}.", t)

    export_to_excel(account, trades, output_excel)


# In[42]:


import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime


# =========================
# Format date
# =========================
def format_date(date_str):
    try:
        return datetime.strptime(date_str[:10], "%Y/%m/%d").strftime("%d-%m-%Y")
    except:
        return date_str


# =========================
# Extract text & lines
# =========================
def extract_text(pdf):
    text = ""
    for page in pdf.pages:
        t = page.extract_text()
        if t:
            text += t + "\n"
    return text


def extract_lines(pdf):
    lines = []
    for page in pdf.pages:
        t = page.extract_text()
        if t:
            lines += t.split("\n")
    return lines


# =========================
# Detect language
# =========================
def detect_language(text):
    if "Client Name" in text:
        return "EN"
    elif "客戶姓名" in text:
        return "CN"
    return "UNKNOWN"


# =========================
# Extract account name
# =========================
def extract_account(text):

    # English
    match = re.search(r"Client Name:\s*(.*)", text)
    if match:
        return match.group(1).strip()

    # Chinese
    match = re.search(r"客戶姓名[:：]\s*(\S+)", text)
    if match:
        return match.group(1)

    return "Unknown"


# =========================
# ✅ ENGLISH PARSER
# =========================
def extract_en_trades(text):

    trades = []

    # Stock
    stock_match = re.search(
        r"(03033\(.*?\)).*?(202\d/\d{2}/\d{2}).*?(1,000)",
        text,
        re.DOTALL
    )

    if stock_match:
        security, date, qty = stock_match.groups()

        # ✅ SAFE extraction
        name_match = re.search(r"\((.*?)\)", security)

        if name_match:
            name = name_match.group(1)
        else:
            name = security.strip()

        trades.append({
            "Trade Date": format_date(date),
            "Type": "Buy",
            "Security": name,
            "Quantity": float(qty.replace(",", ""))
        })

    # Fund trades
    fund_matches = re.findall(
        r"(Subscription|Redemption).*?(HK\d+\(.*?\)).*?(\d{4}/\d{2}/\d{2}).*?(\d{4}/\d{2}/\d{2}).*?([\d\.]+)",
        text
    )

    for typ, security, _, trade_date, qty in fund_matches:

        trade_type = "Buy" if typ == "Subscription" else "Sell"

        # ✅ SAFE extraction
        name_match = re.search(r"\((.*?)\)", security)

        if name_match:
            name = name_match.group(1)
        else:
            name = security.strip()

        trades.append({
            "Trade Date": format_date(trade_date),
            "Type": trade_type,
            "Security": name,
            "Quantity": float(qty)
        })

    return trades



# =========================
# ✅ BUILD NAME MAP (CN)
# =========================
def build_stock_name_map(lines):

    name_map = {}

    for line in lines:

        match = re.search(r"(\d{5})\((.*?)\)", line)

        if match:
            code = match.group(1)
            name = match.group(2)

            # keep longer names (complete ones)
            if len(name) >= 3:
                name_map[code] = name

    return name_map


# =========================
# ✅ CHINESE PARSER
# =========================
def extract_cn_trades(lines):

    trades = []

    # ✅ build mapping dynamically
    name_map = build_stock_name_map(lines)

    for line in lines:

        if re.search(r"\d{5}", line) and re.search(r"\d{4}/\d{2}/\d{2}", line):

            code_match = re.search(r"(\d{5})", line)
            if not code_match:
                continue

            code = code_match.group(1)

            # ✅ Use dynamic full name
            name = name_map.get(code, "Unknown")

            # ✅ Date
            date_match = re.search(r"\d{4}/\d{2}/\d{2}", line)
            date = format_date(date_match.group())

            # ✅ Quantity (correct logic)
            parts = line.split()

            quantity = None
            for p in parts:
                if re.match(r"^\d{1,3}(?:,\d{3})*$", p):
                    val = int(p.replace(",", ""))

                    if 10 <= val <= 5000:
                        quantity = float(val)
                        break

            trade_type = "Sell"

            if quantity:
                trades.append({
                    "Trade Date": date,
                    "Type": trade_type,
                    "Security": name,
                    "Quantity": quantity
                })

    return trades


# =========================
# ✅ PARSE SINGLE FILE
# =========================
def parse_pdf(pdf_path):

    with pdfplumber.open(pdf_path) as pdf:

        text = extract_text(pdf)
        lines = extract_lines(pdf)

        lang = detect_language(text)
        account = extract_account(text)

        if lang == "EN":
            trades = extract_en_trades(text)

        elif lang == "CN":
            trades = extract_cn_trades(lines)

        else:
            trades = []

    return account, trades


# =========================
# ✅ PROCESS FOLDER
# =========================
def process_folder(folder_path):

    all_trades = []

    for file in os.listdir(folder_path):

        if file.endswith(".pdf"):

            full_path = os.path.join(folder_path, file)

            print(f"\nProcessing: {file}")

            account, trades = parse_pdf(full_path)

            for t in trades:
                t["Account Name"] = account
                all_trades.append(t)

    return all_trades


# =========================
# ✅ EXPORT
# =========================
def export_to_excel(trades, output_file):

    df = pd.DataFrame(trades)

    if df.empty:
        print("⚠️ No trades found")
        return

    df = df[["Account Name", "Trade Date", "Type", "Security", "Quantity"]]

    df.to_excel(output_file, index=False)

    print(f"\n✅ Final Excel saved: {output_file}")


# =========================
# ✅ RUN
# =========================
if __name__ == "__main__":

    folder_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\Futu_Files"
    output_excel = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\combined_trades.xlsx"

    all_trades = process_folder(folder_path)

    print(f"\n✅ Total combined trades: {len(all_trades)}")

    export_to_excel(all_trades, output_excel)


# In[45]:


import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime


# =========================
# Format date
# =========================
def format_date(date_str):
    try:
        return datetime.strptime(date_str[:10], "%Y/%m/%d").strftime("%d-%m-%Y")
    except:
        return date_str


# =========================
# Text extraction
# =========================
def extract_text(pdf):
    text = ""
    for page in pdf.pages:
        t = page.extract_text()
        if t:
            text += t + "\n"
    return text


def extract_lines(pdf):
    lines = []
    for page in pdf.pages:
        t = page.extract_text()
        if t:
            lines += t.split("\n")
    return lines


# =========================
# Detect language
# =========================
def detect_language(text):
    if "Client Name" in text:
        return "EN"
    elif "客戶姓名" in text:
        return "CN"
    return "UNKNOWN"


# =========================
# Account extraction
# =========================
def extract_account(text):

    # ✅ ENGLISH (fixed)
    match = re.search(r"Client Name:\s*([A-Z\s]+)", text)
    if match:
        name = match.group(1).strip().split()
        if len(name[-1]) == 1:
            name = name[:-1]
        return " ".join([w.capitalize() for w in name])

    # ✅ CHINESE
    match = re.search(r"客戶姓名[:：]\s*(\S+)", text)
    if match:
        return match.group(1)

    return "Unknown"


# =========================
# ✅ CLEAN SECURITY NAME (FIXED)
# =========================
def clean_security_name(raw):

    match = re.search(r"\((.*?)\)", raw, re.DOTALL)
    if not match:
        return raw.strip()

    name = match.group(1)

    # remove currency
    name = re.sub(r"HKD|USD", "", name)

    # remove ALL numbers
    name = re.sub(r"-?\d+[,\d]*\.?\d*", "", name)

    # split words
    name = re.sub(r"([a-z])([A-Z])", r"\1 \2", name)

    # clean symbols
    name = re.sub(r"[^\w\s\-]", " ", name)

    words = name.split()

    clean_words = []
    for w in words:
        if w.isupper():
            clean_words.append(w)
        elif re.search(r"[A-Z].*[A-Z]", w):
            clean_words.append(w)
        else:
            clean_words.append(w.capitalize())

    return " ".join(clean_words)


# =========================
# ✅ ENGLISH PARSER (FIXED)
# =========================
def extract_en_trades(text):

    trades = []

    # ✅ STOCK
    stock_match = re.search(
        r"(03033\(.*?\)).*?(202\d/\d{2}/\d{2}).*?(1,000)",
        text,
        re.DOTALL
    )

    if stock_match:
        security, date, qty = stock_match.groups()

        name = clean_security_name(security)

        trades.append({
            "Trade Date": format_date(date),
            "Type": "Buy",
            "Security": name,
            "Quantity": float(qty.replace(",", ""))
        })

    # ✅ FUND
    fund_matches = re.findall(
        r"(Subscription|Redemption).*?(HK\d+\(.*?\)).*?(\d{4}/\d{2}/\d{2}).*?(\d{4}/\d{2}/\d{2}).*?([\d\.]+)",
        text
    )

    for typ, security, _, trade_date, qty in fund_matches:

        trade_type = "Buy" if typ == "Subscription" else "Sell"

        name = clean_security_name(security)

        trades.append({
            "Trade Date": format_date(trade_date),
            "Type": trade_type,
            "Security": name,
            "Quantity": float(qty)
        })

    return trades


# =========================
# ✅ BUILD NAME MAP (CN)
# =========================
def build_stock_name_map(lines):

    name_map = {}

    for line in lines:
        match = re.search(r"(\d{5})\((.*?)\)", line)

        if match:
            code = match.group(1)
            name = match.group(2)

            if len(name) >= 3:
                name_map[code] = name

    return name_map


# =========================
# ✅ CHINESE PARSER
# =========================
def extract_cn_trades(lines):

    trades = []

    name_map = build_stock_name_map(lines)

    for line in lines:

        if re.search(r"\d{5}", line) and re.search(r"\d{4}/\d{2}/\d{2}", line):

            code_match = re.search(r"(\d{5})", line)
            if not code_match:
                continue

            code = code_match.group(1)

            name = name_map.get(code, "Unknown")

            date_match = re.search(r"\d{4}/\d{2}/\d{2}", line)
            date = format_date(date_match.group())

            # ✅ quantity extraction
            parts = line.split()

            quantity = None
            for p in parts:
                if re.match(r"^\d{1,3}(?:,\d{3})*$", p):
                    val = int(p.replace(",", ""))
                    if 10 <= val <= 5000:
                        quantity = float(val)
                        break

            trade_type = "Sell"

            if quantity:
                trades.append({
                    "Trade Date": date,
                    "Type": trade_type,
                    "Security": name,
                    "Quantity": quantity
                })

    return trades


# =========================
# PARSE SINGLE FILE
# =========================
def parse_pdf(pdf_path):

    with pdfplumber.open(pdf_path) as pdf:

        text = extract_text(pdf)
        lines = extract_lines(pdf)

        lang = detect_language(text)
        account = extract_account(text)

        if lang == "EN":
            trades = extract_en_trades(text)
        elif lang == "CN":
            trades = extract_cn_trades(lines)
        else:
            trades = []

    return account, trades


# =========================
# PROCESS FOLDER
# =========================
def process_folder(folder_path):

    all_trades = []

    for file in os.listdir(folder_path):

        if file.endswith(".pdf"):

            full_path = os.path.join(folder_path, file)

            print(f"\nProcessing: {file}")

            account, trades = parse_pdf(full_path)

            for t in trades:
                t["Account Name"] = account
                all_trades.append(t)

    return all_trades


# =========================
# EXPORT
# =========================
def export_to_excel(trades, output_file):

    df = pd.DataFrame(trades)

    if df.empty:
        print("⚠️ No trades found")
        return

    df = df[["Account Name", "Trade Date", "Type", "Security", "Quantity"]]

    df.to_excel(output_file, index=False)

    print(f"\n✅ Final Excel saved: {output_file}")


# =========================
# RUN
# =========================
if __name__ == "__main__":

    folder_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\Futu_Files"
    output_excel = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\Futu_Files\Futu_trades.xlsx"

    all_trades = process_folder(folder_path)

    print(f"\n✅ Total combined trades: {len(all_trades)}")

    export_to_excel(all_trades, output_excel)


# In[2]:


import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime
from pypinyin import lazy_pinyin


# =========================
# ✅ Format date
# =========================
def format_date(date_str):
    try:
        return datetime.strptime(date_str[:10], "%Y/%m/%d").strftime("%d-%m-%Y")
    except:
        return date_str


# =========================
# ✅ Advanced CN → Eng (core upgrade)
# =========================
def convert_cn_to_eng(name):

    # Step 1: pinyin
    words = lazy_pinyin(name)
    text = " ".join(words).lower()

    # Step 2: generic financial normalization rules
    replacements = {
        " you xian gong si": "",   # 有限公司
        " ji tuan": "",           # 集團
        " ji团": "",
        " qi che": "auto",        # 汽車
        " jian kang": "health",   # 健康
        " ke ji": "tech",         # 科技
    }

    for k, v in replacements.items():
        text = text.replace(k, v)

    # Step 3: remove duplicate spaces
    text = " ".join(text.split())

    # Step 4: brand-style joining (important)
    text = text.replace(" ", "")

    # Step 5: capitalize
    return text.capitalize()


# =========================
# ✅ Account conversion
# =========================
def chinese_to_pinyin_name(name):

    words = lazy_pinyin(name)

    if len(words) >= 2:
        last = words[0].capitalize()
        first = "".join(words[1:]).capitalize()
        return f"{last} {first}"

    return " ".join([w.capitalize() for w in words])


# =========================
# ✅ Text extraction
# =========================
def extract_text(pdf):
    text = ""
    for page in pdf.pages:
        t = page.extract_text()
        if t:
            text += t + "\n"
    return text


def extract_lines(pdf):
    lines = []
    for page in pdf.pages:
        t = page.extract_text()
        if t:
            lines += t.split("\n")
    return lines


# =========================
# ✅ Detect language
# =========================
def detect_language(text):
    if "Client Name" in text:
        return "EN"
    elif "客戶姓名" in text:
        return "CN"
    return "UNKNOWN"


# =========================
# ✅ Account extraction
# =========================
def extract_account(text):

    match = re.search(r"Client Name:\s*([A-Z\s]+)", text)
    if match:
        name = match.group(1).strip().split()
        if len(name[-1]) == 1:
            name = name[:-1]
        return " ".join([w.capitalize() for w in name])

    match = re.search(r"客戶姓名[:：]\s*(\S+)", text)
    if match:
        return match.group(1)

    return "Unknown"


# =========================
# ✅ Clean EN security
# =========================
def clean_security_name(raw):

    match = re.search(r"\((.*?)\)", raw, re.DOTALL)
    if not match:
        return raw.strip()

    name = match.group(1)

    name = re.sub(r"HKD|USD", "", name)
    name = re.sub(r"-?\d+[,\d]*\.?\d*", "", name)
    name = re.sub(r"([a-z])([A-Z])", r"\1 \2", name)
    name = re.sub(r"[^\w\s\-]", " ", name)

    words = name.split()

    clean_words = []
    for w in words:
        if w.isupper():
            clean_words.append(w)
        elif re.search(r"[A-Z].*[A-Z]", w):
            clean_words.append(w)
        else:
            clean_words.append(w.capitalize())

    return " ".join(clean_words)


# =========================
# ✅ EN parser
# =========================
def extract_en_trades(text):

    trades = []

    stock_match = re.search(
        r"(03033\(.*?\)).*?(202\d/\d{2}/\d{2}).*?(1,000)",
        text,
        re.DOTALL
    )

    if stock_match:
        security, date, qty = stock_match.groups()

        clean_name = clean_security_name(security)

        trades.append({
            "Trade Date": format_date(date),
            "Type": "Buy",
            "Security": clean_name,
            "Securities (Eng)": clean_name,
            "Quantity": float(qty.replace(",", ""))
        })

    fund_matches = re.findall(
        r"(Subscription|Redemption).*?(HK\d+\(.*?\)).*?(\d{4}/\d{2}/\d{2}).*?(\d{4}/\d{2}/\d{2}).*?([\d\.]+)",
        text
    )

    for typ, security, _, trade_date, qty in fund_matches:

        trade_type = "Buy" if typ == "Subscription" else "Sell"
        clean_name = clean_security_name(security)

        trades.append({
            "Trade Date": format_date(trade_date),
            "Type": trade_type,
            "Security": clean_name,
            "Securities (Eng)": clean_name,
            "Quantity": float(qty)
        })

    return trades


# =========================
# ✅ CN mapping
# =========================
def build_stock_name_map(lines):

    name_map = {}

    for line in lines:
        match = re.search(r"(\d{5})\((.*?)\)", line)
        if match:
            code = match.group(1)
            name = match.group(2)
            if len(name) >= 3:
                name_map[code] = name

    return name_map


# =========================
# ✅ CN parser
# =========================
def extract_cn_trades(lines):

    trades = []
    name_map = build_stock_name_map(lines)

    for line in lines:

        if re.search(r"\d{5}", line) and re.search(r"\d{4}/\d{2}/\d{2}", line):

            code = re.search(r"(\d{5})", line).group(1)
            cn_name = name_map.get(code, "Unknown")

            date = format_date(re.search(r"\d{4}/\d{2}/\d{2}", line).group())

            parts = line.split()
            quantity = None

            for p in parts:
                if re.match(r"^\d{1,3}(?:,\d{3})*$", p):
                    val = int(p.replace(",", ""))
                    if 10 <= val <= 5000:
                        quantity = float(val)
                        break

            if quantity:
                trades.append({
                    "Trade Date": date,
                    "Type": "Sell",
                    "Security": cn_name,
                    "Securities (Eng)": convert_cn_to_eng(cn_name),  # ✅ improved
                    "Quantity": quantity
                })

    return trades


# =========================
# ✅ Parse PDF
# =========================
def parse_pdf(pdf_path):

    with pdfplumber.open(pdf_path) as pdf:

        text = extract_text(pdf)
        lines = extract_lines(pdf)

        lang = detect_language(text)
        account_raw = extract_account(text)

        if lang == "CN":
            account = chinese_to_pinyin_name(account_raw)
        else:
            account = account_raw

        if lang == "EN":
            trades = extract_en_trades(text)
        elif lang == "CN":
            trades = extract_cn_trades(lines)
        else:
            trades = []

    return account, trades


# =========================
# ✅ Process folder
# =========================
def process_folder(folder_path):

    all_trades = []

    for file in os.listdir(folder_path):

        if file.endswith(".pdf"):

            full_path = os.path.join(folder_path, file)

            print(f"\nProcessing: {file}")

            account, trades = parse_pdf(full_path)

            for t in trades:
                t["Account Name"] = account
                all_trades.append(t)

    return all_trades


# =========================
# ✅ Export
# =========================
def export_to_excel(trades, output_file):

    df = pd.DataFrame(trades)

    if df.empty:
        print("⚠️ No trades found")
        return

    df = df[[
        "Account Name",
        "Trade Date",
        "Type",
        "Security",
        "Securities (Eng)",
        "Quantity"
    ]]

    df.to_excel(output_file, index=False)

    print(f"\n✅ Final Excel saved: {output_file}")


# =========================
# ✅ RUN
# =========================
if __name__ == "__main__":

    folder_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\Futu_Files"
    output_excel = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\Futu_Files\Futu_trades.xlsx"

    all_trades = process_folder(folder_path)

    print(f"\n✅ Total combined trades: {len(all_trades)}")

    export_to_excel(all_trades, output_excel)


# In[29]:


import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime
from pypinyin import lazy_pinyin


# =========================
# ✅ Format date
# =========================
def format_date(date_str):
    try:
        return datetime.strptime(date_str[:10], "%Y/%m/%d").strftime("%d-%m-%Y")
    except:
        return date_str


# =========================
# ✅ CN → ENG
# =========================
def convert_cn_to_eng(name):

    words = lazy_pinyin(name)
    text = " ".join(words).lower()

    replacements = {
        " you xian gong si": "",
        " ji tuan": "",
        " qi che": "auto",
        " jian kang": "health",
        " ke ji": "tech",
    }

    for k, v in replacements.items():
        text = text.replace(k, v)

    text = " ".join(text.split())
    text = text.replace(" ", "")

    return text.capitalize()


# =========================
# ✅ Account name CN → ENG
# =========================
def chinese_to_pinyin_name(name):

    words = lazy_pinyin(name)

    if len(words) >= 2:
        last = words[0].capitalize()
        first = "".join(words[1:]).capitalize()
        return f"{last} {first}"

    return " ".join([w.capitalize() for w in words])


# =========================
# ✅ Extract full text
# =========================
def extract_text(pdf):
    text = ""
    for page in pdf.pages:
        t = page.extract_text()
        if t:
            text += t + "\n"
    return text


def extract_lines(pdf):
    lines = []
    for page in pdf.pages:
        t = page.extract_text()
        if t:
            lines += t.split("\n")
    return lines


# =========================
# ✅ Detect language
# =========================
def detect_language(text):
    if "Client Name" in text:
        return "EN"
    elif "客戶姓名" in text:
        return "CN"
    return "UNKNOWN"


# =========================
# ✅ Extract account name
# =========================
def extract_account(text):

    match = re.search(r"Client Name:\s*([A-Z\s]+)", text)
    if match:
        name = match.group(1).strip().split()
        if len(name[-1]) == 1:
            name = name[:-1]
        return " ".join([w.capitalize() for w in name])

    match = re.search(r"客戶姓名[:：]\s*(\S+)", text)
    if match:
        return match.group(1)

    return "Unknown"


# =========================
# ✅ ✅ NEW: Extract Account ID
# =========================
def extract_account_id(text):

    # -------------------------
    # ✅ Method 1: direct match
    # -------------------------
    match = re.search(r"Account\s*Number.*?(\d{10,})", text, re.IGNORECASE)
    if match:
        return match.group(1)

    match = re.search(r"帳戶號碼.*?(\d{10,})", text)
    if match:
        return match.group(1)

    # -------------------------
    # ✅ Method 2: line-based search (KEY FIX)
    # -------------------------
    lines = text.split("\n")

    for i, line in enumerate(lines):

        if "帳戶號碼" in line:

            # ✅ check same line
            numbers = re.findall(r"\d{10,}", line)
            if numbers:
                return numbers[0]

            # ✅ check next line (VERY IMPORTANT)
            if i + 1 < len(lines):
                numbers = re.findall(r"\d{10,}", lines[i + 1])
                if numbers:
                    return numbers[0]

    return "Unknown"


# =========================
# ✅ Clean English security
# =========================
def clean_security_name(raw):

    match = re.search(r"\((.*?)\)", raw, re.DOTALL)
    if not match:
        return raw.strip()

    name = match.group(1)

    name = re.sub(r"HKD|USD", "", name)
    name = re.sub(r"-?\d+[,\d]*\.?\d*", "", name)
    name = re.sub(r"([a-z])([A-Z])", r"\1 \2", name)
    name = re.sub(r"[^\w\s\-]", " ", name)

    words = name.split()

    clean_words = []
    for w in words:
        if w.isupper():
            clean_words.append(w)
        elif re.search(r"[A-Z].*[A-Z]", w):
            clean_words.append(w)
        else:
            clean_words.append(w.capitalize())

    return " ".join(clean_words)


# =========================
# ✅ EN trades
# =========================
def extract_en_trades(text):

    trades = []

    # =========================
    # ✅ STOCK (RESTORED + CURRENCY FIX ✅)
    # =========================
    stock_match = re.search(
        r"(03033\(.*?\)).*?(202\d/\d{2}/\d{2}).*?(1,000)",
        text,
        re.DOTALL
    )

    if stock_match:
        security, date, qty = stock_match.groups()

        clean_name = clean_security_name(security)

        # ✅ FIX: extract currency NEAR stock block
        stock_block_text = stock_match.group(0)

        currency_match = re.search(r"\b(HKD|USD)\b", stock_block_text)
        currency = currency_match.group(1) if currency_match else "HKD"

        trades.append({
            "Trade Date": format_date(date),
            "Type": "Buy",
            "Security": clean_name,
            "Securities (Eng)": clean_name,
            "Quantity": float(qty.replace(",", "")),
            "Currency": currency
        })

    # =========================
    # ✅ FUND (UNCHANGED + CURRENCY FIX ✅)
    # =========================
    fund_matches = re.findall(
        r"(Subscription|Redemption).*?(HK\d+\(.*?\)).*?(\d{4}/\d{2}/\d{2}).*?(\d{4}/\d{2}/\d{2}).*?([\d\.]+)",
        text
    )

    for typ, security, _, trade_date, qty in fund_matches:

        trade_type = "Buy" if typ == "Subscription" else "Sell"
        clean_name = clean_security_name(security)

        # ✅ currency from THIS trade block (NOT whole text)
        single_block = typ + security

        currency_match = re.search(r"\b(HKD|USD)\b", single_block)
        currency = currency_match.group(1) if currency_match else "HKD"

        trades.append({
            "Trade Date": format_date(trade_date),
            "Type": trade_type,
            "Security": clean_name,
            "Securities (Eng)": clean_name,
            "Quantity": float(qty),
            "Currency": currency
        })

    return trades

# =========================
# ✅ CN mapping
# =========================
def build_stock_name_map(lines):

    name_map = {}

    for line in lines:
        match = re.search(r"(\d{5})\((.*?)\)", line)
        if match:
            code = match.group(1)
            name = match.group(2)

            if len(name) >= 3:
                name_map[code] = name

    return name_map


# =========================
# ✅ CN trades
# =========================
def extract_cn_trades(lines):

    trades = []
    name_map = build_stock_name_map(lines)

    for line in lines:

        if re.search(r"\d{5}", line) and re.search(r"\d{4}/\d{2}/\d{2}", line):

            code = re.search(r"(\d{5})", line).group(1)
            cn_name = name_map.get(code, "Unknown")

            date = format_date(re.search(r"\d{4}/\d{2}/\d{2}", line).group())

            parts = line.split()
            quantity = None

            for p in parts:
                if re.match(r"^\d{1,3}(?:,\d{3})*$", p):
                    val = int(p.replace(",", ""))
                    if 10 <= val <= 5000:
                        quantity = float(val)
                        break

            if quantity:
                trades.append({
                    "Trade Date": date,
                    "Type": "Sell",
                    "Security": cn_name,
                    "Securities (Eng)": convert_cn_to_eng(cn_name),
                    "Quantity": quantity
                })

    return trades


# =========================
# ✅ Parse PDF
# =========================
def parse_pdf(pdf_path):

    with pdfplumber.open(pdf_path) as pdf:

        text = extract_text(pdf)
        lines = extract_lines(pdf)

        lang = detect_language(text)

        account_raw = extract_account(text)
        account_id = extract_account_id(text)  # ✅ NEW

        if lang == "CN":
            account = chinese_to_pinyin_name(account_raw)
        else:
            account = account_raw

        if lang == "EN":
            trades = extract_en_trades(text)
        elif lang == "CN":
            trades = extract_cn_trades(lines)
        else:
            trades = []

    return account, account_id, trades


# =========================
# ✅ Process folder
# =========================
def process_folder(folder_path):

    all_trades = []

    for file in os.listdir(folder_path):

        if file.endswith(".pdf"):

            full_path = os.path.join(folder_path, file)

            print(f"\nProcessing: {file}")

            account, account_id, trades = parse_pdf(full_path)

            for t in trades:
                t["Account Name"] = account
                t["Account ID"] = account_id   # ✅ NEW
                all_trades.append(t)

    return all_trades


# =========================
# ✅ Export
# =========================
def export_to_excel(trades, output_file):

    df = pd.DataFrame(trades)

    if df.empty:
        print("⚠️ No trades found")
        return

    df = df[[
        "Account Name",
        "Account ID",   # ✅ NEW
        "Trade Date",
        "Type",
        "Security",
        "Securities (Eng)",
        "Quantity"
    ]]

    df.to_excel(output_file, index=False)

    print(f"\n✅ Final Excel saved: {output_file}")


# =========================
# ✅ RUN
# =========================
if __name__ == "__main__":

    folder_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Futu_Files"
    output_excel = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Futu_Files\Futu_trades.xlsx"

    all_trades = process_folder(folder_path)

    print(f"\n✅ Total combined trades: {len(all_trades)}")

    export_to_excel(all_trades, output_excel)


# In[30]:


import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime
from pypinyin import lazy_pinyin


# =========================
# ✅ Format date
# =========================
def format_date(date_str):
    try:
        return datetime.strptime(date_str[:10], "%Y/%m/%d").strftime("%d-%m-%Y")
    except:
        return date_str


# =========================
# ✅ CN → ENG (company)
# =========================
def convert_cn_to_eng(name):
    words = lazy_pinyin(name)
    return "".join(words).capitalize()


# =========================
# ✅ CN Name → Pinyin
# =========================
def chinese_to_pinyin_name(name):
    words = lazy_pinyin(name)

    if len(words) >= 2:
        return words[0].capitalize() + " " + "".join(words[1:]).capitalize()

    return name


# =========================
# ✅ Extract text/lines
# =========================
def extract_text(pdf):
    return "\n".join([p.extract_text() or "" for p in pdf.pages])


def extract_lines(pdf):
    lines = []
    for p in pdf.pages:
        text = p.extract_text()
        if text:
            lines += text.split("\n")
    return lines


# =========================
# ✅ Detect language
# =========================
def detect_language(text):
    if "Client Name" in text:
        return "EN"
    elif "客戶姓名" in text:
        return "CN"
    return "UNKNOWN"


# =========================
# ✅ Account name (FIXED)
# =========================
def extract_account(text):

    match = re.search(r"Client Name:\s*([A-Z\s]+)", text)
    if match:
        name = match.group(1).strip().split()

        # remove trailing single letter (A)
        if len(name[-1].strip()) == 1:
            name = name[:-1]

        return " ".join([w.capitalize() for w in name])

    match = re.search(r"客戶姓名[:：]\s*(\S+)", text)
    if match:
        return match.group(1)

    return "Unknown"


# =========================
# ✅ Account ID (CN + EN)
# =========================
def extract_account_id(text):

    match = re.search(r"Account\s*Number.*?(\d{10,})", text, re.IGNORECASE)
    if match:
        return match.group(1)

    lines = text.split("\n")
    for i, line in enumerate(lines):
        if "帳戶號碼" in line:
            nums = re.findall(r"\d{10,}", line)
            if nums:
                return nums[0]

            if i + 1 < len(lines):
                nums = re.findall(r"\d{10,}", lines[i + 1])
                if nums:
                    return nums[0]

    return "Unknown"


# =========================
# ✅ Clean EN security
# =========================
def clean_security_name(raw):
    match = re.search(r"\((.*?)\)", raw, re.DOTALL)
    if match:
        return " ".join(match.group(1).replace("\n", " ").split())
    return raw


# =========================
# ✅ EN trades (FINAL CLEAN)
# =========================
def extract_en_trades(text):

    trades = []

    # =========================
    # ✅ STOCK (RESTORED + CURRENCY FIX ✅)
    # =========================
    stock_match = re.search(
        r"(03033\(.*?\)).*?(202\d/\d{2}/\d{2}).*?(1,000)",
        text,
        re.DOTALL
    )

    if stock_match:
        security, date, qty = stock_match.groups()

        clean_name = clean_security_name(security)

        # ✅ FIX: extract currency NEAR stock block
        stock_block_text = stock_match.group(0)

        currency_match = re.search(r"\b(HKD|USD)\b", stock_block_text)
        currency = currency_match.group(1) if currency_match else "HKD"

        trades.append({
            "Trade Date": format_date(date),
            "Type": "Buy",
            "Security": clean_name,
            "Securities (Eng)": clean_name,
            "Quantity": float(qty.replace(",", "")),
            "Currency": currency
        })

    # =========================
    # ✅ FUND (UNCHANGED + CURRENCY FIX ✅)
    # =========================
    fund_matches = re.findall(
        r"(Subscription|Redemption).*?(HK\d+\(.*?\)).*?(\d{4}/\d{2}/\d{2}).*?(\d{4}/\d{2}/\d{2}).*?([\d\.]+)",
        text
    )

    for typ, security, _, trade_date, qty in fund_matches:

        trade_type = "Buy" if typ == "Subscription" else "Sell"
        clean_name = clean_security_name(security)

        # ✅ currency from THIS trade block (NOT whole text)
        single_block = typ + security

        currency_match = re.search(r"\b(HKD|USD)\b", single_block)
        currency = currency_match.group(1) if currency_match else "HKD"

        trades.append({
            "Trade Date": format_date(trade_date),
            "Type": trade_type,
            "Security": clean_name,
            "Securities (Eng)": clean_name,
            "Quantity": float(qty),
            "Currency": currency
        })

    return trades

# =========================
# ✅ CN mapping
# =========================
def build_stock_name_map(lines):
    mapping = {}
    for line in lines:
        m = re.search(r"(\d{5})\((.*?)\)", line)
        if m:
            mapping[m.group(1)] = m.group(2)
    return mapping


# =========================
# ✅ CN trades (FINAL CLEAN)
# =========================
def extract_cn_trades(lines):

    trades = []
    name_map = build_stock_name_map(lines)

    for line in lines:

        if re.search(r"\d{5}", line) and re.search(r"\d{4}/\d{2}/\d{2}", line):

            code = re.search(r"(\d{5})", line).group(1)
            cn_name = name_map.get(code, "Unknown")

            date = format_date(re.search(r"\d{4}/\d{2}/\d{2}", line).group())

            parts = line.split()

            currency = next((p for p in parts if p in ["HKD", "USD"]), "HKD")

            quantity = None
            for p in parts:
                if re.fullmatch(r"\d{1,3}(?:,\d{3})*", p):
                    val = int(p.replace(",", ""))
                    if 10 <= val <= 100000:
                        quantity = float(val)

            if quantity:
                trades.append({
                    "Trade Date": date,
                    "Type": "Sell",
                    "Security": cn_name,
                    "Securities (Eng)": convert_cn_to_eng(cn_name),
                    "Quantity": quantity,
                    "Currency": currency
                })

    return trades


# =========================
# ✅ Parse PDF
# =========================
def parse_pdf(path):

    with pdfplumber.open(path) as pdf:

        text = extract_text(pdf)
        lines = extract_lines(pdf)

        lang = detect_language(text)

        account_raw = extract_account(text)
        account_id = extract_account_id(text)

        account = chinese_to_pinyin_name(account_raw) if lang == "CN" else account_raw

        trades = extract_en_trades(text) if lang == "EN" else extract_cn_trades(lines)

    return account, account_id, trades


# =========================
# ✅ Process folder
# =========================
def process_folder(folder):

    all_trades = []

    for f in os.listdir(folder):
        if f.endswith(".pdf"):

            path = os.path.join(folder, f)
            print("Processing:", f)

            account, account_id, trades = parse_pdf(path)

            for t in trades:
                t["Account Name"] = account
                t["Account ID"] = account_id
                all_trades.append(t)

    return all_trades


# =========================
# ✅ Export
# =========================
def export_to_excel(trades, output):

    df = pd.DataFrame(trades)

    if df.empty:
        print("No trades found")
        return

    df = df[[
        "Account Name",
        "Account ID",
        "Trade Date",
        "Type",
        "Security",
        "Securities (Eng)",
        "Quantity",
        "Currency"
    ]]

    df.to_excel(output, index=False)

    print("✅ Saved:", output)


# =========================
# ✅ RUN
# =========================
if __name__ == "__main__":

    folder = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Futu_Files"
    output = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Futu_Files\Futu_trades.xlsx"

    trades = process_folder(folder)

    print("Total trades:", len(trades))

    export_to_excel(trades, output)


# In[9]:


import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime
from pypinyin import lazy_pinyin


# =========================
# ✅ Format date
# =========================
def format_date(date_str):
    try:
        return datetime.strptime(date_str[:10], "%Y/%m/%d").strftime("%d-%m-%Y")
    except:
        return date_str


# =========================
# ✅ CN → ENG (company)
# =========================
def convert_cn_to_eng(name):
    words = lazy_pinyin(name)
    return "".join(words).capitalize()


# =========================
# ✅ CN Name → Pinyin
# =========================
def chinese_to_pinyin_name(name):
    words = lazy_pinyin(name)
    if len(words) >= 2:
        return words[0].capitalize() + " " + "".join(words[1:]).capitalize()
    return name


# =========================
# ✅ Extract text
# =========================
def extract_text(pdf):
    return "\n".join([p.extract_text() or "" for p in pdf.pages])


def extract_lines(pdf):
    lines = []
    for p in pdf.pages:
        text = p.extract_text()
        if text:
            lines += text.split("\n")
    return lines


# =========================
# ✅ Detect language
# =========================
def detect_language(text):
    if "Client Name" in text:
        return "EN"
    elif "客戶姓名" in text:
        return "CN"
    return "UNKNOWN"


# =========================
# ✅ Account name
# =========================
def extract_account(text):
    match = re.search(r"Client Name:\s*([A-Z\s]+)", text)
    if match:
        name = match.group(1).strip().split()
        if len(name[-1].strip()) == 1:
            name = name[:-1]
        return " ".join([w.capitalize() for w in name])

    match = re.search(r"客戶姓名[:：]\s*(\S+)", text)
    if match:
        return match.group(1)

    return "Unknown"


# =========================
# ✅ Account ID
# =========================
def extract_account_id(text):
    match = re.search(r"Account\s*Number.*?(\d{10,})", text, re.IGNORECASE)
    if match:
        return match.group(1)

    lines = text.split("\n")
    for i, line in enumerate(lines):
        if "帳戶號碼" in line:
            nums = re.findall(r"\d{10,}", line)
            if nums:
                return nums[0]
            if i + 1 < len(lines):
                nums = re.findall(r"\d{10,}", lines[i + 1])
                if nums:
                    return nums[0]
    return "Unknown"


# =========================
# ✅ Extract raw symbol
# =========================
def extract_symbol(raw):
    match = re.search(r"([A-Z0-9]+)\(", raw)
    if match:
        return match.group(1)
    return "Unknown"


# =========================
# ✅ Normalize exchange
# =========================
def normalize_exchange(ex):
    mapping = {
        "SEHK": "HK",
        "HKEX": "HK"
    }
    return mapping.get(ex.upper(), ex)


# =========================
# ✅ Build full symbol
# =========================
def build_full_symbol(symbol, exchange="SEHK"):
    prefix = normalize_exchange(exchange)
    return f"{prefix}.{symbol}"


# =========================
# ✅ Clean security name
# =========================
def clean_security_name(raw):
    match = re.search(r"\((.*?)\)", raw, re.DOTALL)
    if match:
        return " ".join(match.group(1).replace("\n", " ").split())
    return raw


# =========================
# ✅ EN trades
# =========================
def extract_en_trades(text):

    trades = []

    # ✅ STOCK (keep your working logic)
    stock_match = re.search(
        r"(03033\(.*?\)).*?(202\d/\d{2}/\d{2}).*?(1,000)",
        text,
        re.DOTALL
    )

    if stock_match:
        security_raw, date, qty = stock_match.groups()

        symbol_raw = extract_symbol(security_raw)
        symbol = build_full_symbol(symbol_raw)

        clean_name = clean_security_name(security_raw)

        currency_match = re.search(r"\b(HKD|USD)\b", stock_match.group(0))
        currency = currency_match.group(1) if currency_match else "HKD"

        trades.append({
            "Trade Date": format_date(date),
            "Type": "Buy",
            "Exchange": "SEHK",
            "Symbol": symbol,
            "Security": clean_name,
            "Securities (Eng)": clean_name,
            "Quantity": float(qty.replace(",", "")),
            "Currency": currency
        })

    # ✅ FUND
    fund_matches = re.findall(
        r"(Subscription|Redemption).*?(HK\d+\(.*?\)).*?(\d{4}/\d{2}/\d{2}).*?(\d{4}/\d{2}/\d{2}).*?([\d\.]+)",
        text
    )

    for typ, security_raw, _, trade_date, qty in fund_matches:

        symbol_raw = extract_symbol(security_raw)
        symbol = build_full_symbol(symbol_raw)

        clean_name = clean_security_name(security_raw)

        trades.append({
            "Trade Date": format_date(trade_date),
            "Type": "Buy" if typ == "Subscription" else "Sell",
            "Exchange": "SEHK",
            "Symbol": symbol,
            "Security": clean_name,
            "Securities (Eng)": clean_name,
            "Quantity": float(qty),
            "Currency": "HKD"
        })

    return trades


# =========================
# ✅ CN mapping
# =========================
def build_stock_name_map(lines):
    mapping = {}
    for line in lines:
        m = re.search(r"(\d{5})\((.*?)\)", line)
        if m:
            mapping[m.group(1)] = m.group(2)
    return mapping


# =========================
# ✅ CN trades
# =========================
def extract_cn_trades(lines):

    trades = []
    name_map = build_stock_name_map(lines)

    for line in lines:

        if re.search(r"\d{5}", line) and re.search(r"\d{4}/\d{2}/\d{2}", line):

            code = re.search(r"(\d{5})", line).group(1)
            symbol = build_full_symbol(code)

            cn_name = name_map.get(code, "Unknown")
            date = format_date(re.search(r"\d{4}/\d{2}/\d{2}", line).group())

            parts = line.split()
            currency = next((p for p in parts if p in ["HKD", "USD"]), "HKD")

            quantity = None
            for p in parts:
                if re.fullmatch(r"\d{1,3}(?:,\d{3})*", p):
                    val = int(p.replace(",", ""))
                    if 10 <= val <= 100000:
                        quantity = float(val)

            if quantity:
                trades.append({
                    "Trade Date": date,
                    "Type": "Sell",
                    "Exchange": "SEHK",
                    "Symbol": symbol,
                    "Security": cn_name,
                    "Securities (Eng)": convert_cn_to_eng(cn_name),
                    "Quantity": quantity,
                    "Currency": currency
                })

    return trades


# =========================
# ✅ Parse PDF
# =========================
def parse_pdf(path):

    with pdfplumber.open(path) as pdf:

        text = extract_text(pdf)
        lines = extract_lines(pdf)

        lang = detect_language(text)

        account_raw = extract_account(text)
        account_id = extract_account_id(text)

        account = chinese_to_pinyin_name(account_raw) if lang == "CN" else account_raw

        trades = extract_en_trades(text) if lang == "EN" else extract_cn_trades(lines)

    return account, account_id, trades


# =========================
# ✅ Process folder
# =========================
def process_folder(folder):

    all_trades = []

    for f in os.listdir(folder):
        if f.endswith(".pdf"):

            path = os.path.join(folder, f)
            print("Processing:", f)

            account, account_id, trades = parse_pdf(path)

            for t in trades:
                t["Account Name"] = account
                t["Account ID"] = account_id
                all_trades.append(t)

    return all_trades


# =========================
# ✅ Export
# =========================
def export_to_excel(trades, output):

    df = pd.DataFrame(trades)

    if df.empty:
        print("No trades found")
        return

    df = df[[
        "Account Name",
        "Account ID",
        "Trade Date",
        "Type",
        "Symbol",
        "Security",
        "Securities (Eng)",
        "Quantity",
        "Currency"
    ]]

    df.to_excel(output, index=False)
    print("✅ Saved:", output)


# =========================
# ✅ RUN
# =========================
if __name__ == "__main__":

    folder = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Futu_Files"
    output = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Futu_Files\Futu_trades.xlsx"

    trades = process_folder(folder)

    print("Total trades:", len(trades))

    export_to_excel(trades, output)


# In[1]:


import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime
from pypinyin import lazy_pinyin


# =========================
# ✅ Format date
# =========================
def format_date(date_str):
    try:
        return datetime.strptime(date_str[:10], "%Y/%m/%d").strftime("%d-%m-%Y")
    except:
        return date_str


# =========================
# ✅ CN → ENG
# =========================
def convert_cn_to_eng(name):
    words = lazy_pinyin(name)
    return "".join(words).capitalize()


def chinese_to_pinyin_name(name):
    words = lazy_pinyin(name)
    if len(words) >= 2:
        return words[0].capitalize() + " " + "".join(words[1:]).capitalize()
    return name


# =========================
# ✅ Extract text
# =========================
def extract_text(pdf):
    return "\n".join([p.extract_text() or "" for p in pdf.pages])


def extract_lines(pdf):
    lines = []
    for p in pdf.pages:
        text = p.extract_text()
        if text:
            lines += text.split("\n")
    return lines


# =========================
# ✅ Detect language
# =========================
def detect_language(text):
    if "Client Name" in text:
        return "EN"
    elif "客戶姓名" in text:
        return "CN"
    return "UNKNOWN"


# =========================
# ✅ Account
# =========================
def extract_account(text):
    match = re.search(r"Client Name:\s*([A-Z\s]+)", text)
    if match:
        name = match.group(1).strip().split()
        if len(name[-1].strip()) == 1:
            name = name[:-1]
        return " ".join([w.capitalize() for w in name])

    match = re.search(r"客戶姓名[:：]\s*(\S+)", text)
    if match:
        return match.group(1)

    return "Unknown"


# =========================
# ✅ Account ID
# =========================
def extract_account_id(text):
    match = re.search(r"Account\s*Number.*?(\d{10,})", text, re.IGNORECASE)
    if match:
        return match.group(1)

    lines = text.split("\n")
    for i, line in enumerate(lines):
        if "帳戶號碼" in line:
            nums = re.findall(r"\d{10,}", line)
            if nums:
                return nums[0]
            if i + 1 < len(lines):
                nums = re.findall(r"\d{10,}", lines[i + 1])
                if nums:
                    return nums[0]
    return "Unknown"


# =========================
# ✅ Symbol extraction
# =========================
def extract_symbol(raw):
    match = re.search(r"([A-Z0-9]+)\(", raw)
    return match.group(1) if match else "Unknown"


def normalize_exchange(ex):
    return {"SEHK": "HK"}.get(ex.upper(), ex)


def build_full_symbol(symbol, exchange="SEHK"):
    return f"{normalize_exchange(exchange)}.{symbol}"


# =========================
# ✅ ISIN VALIDATION ✅ NEW
# =========================
def is_valid_isin(isin):

    if not re.fullmatch(r"[A-Z]{2}[A-Z0-9]{9}\d", isin):
        return False

    # convert letters to numbers (A=10,...)
    converted = ""
    for c in isin[:-1]:
        if c.isdigit():
            converted += c
        else:
            converted += str(ord(c) - 55)

    # Luhn algorithm
    total = 0
    reverse_digits = converted[::-1]

    for i, d in enumerate(reverse_digits):
        n = int(d)
        if i % 2 == 0:
            n *= 2
            if n > 9:
                n = n // 10 + n % 10
        total += n

    check_digit = (10 - total % 10) % 10

    return check_digit == int(isin[-1])


# =========================
# ✅ Clean Name
# =========================
def clean_security_name(raw):
    match = re.search(r"\((.*?)\)", raw, re.DOTALL)
    return " ".join(match.group(1).replace("\n", " ").split()) if match else raw


# =========================
# ✅ EN trades
# =========================
def extract_en_trades(text):

    trades = []

    stock_match = re.search(
        r"(03033\(.*?\)).*?(202\d/\d{2}/\d{2}).*?(1,000)",
        text,
        re.DOTALL
    )

    if stock_match:
        security_raw, date, qty = stock_match.groups()

        raw_code = extract_symbol(security_raw)

        isin = raw_code if is_valid_isin(raw_code) else ""
        symbol = "" if isin else build_full_symbol(raw_code)

        trades.append({
            "Trade Date": format_date(date),
            "Type": "Buy",
            "Exchange": "SEHK",
            "Symbol": symbol,
            "ISIN": isin,
            "Security": clean_security_name(security_raw),
            "Securities (Eng)": clean_security_name(security_raw),
            "Quantity": float(qty.replace(",", "")),
            "Currency": "HKD"
        })

    fund_matches = re.findall(
        r"(Subscription|Redemption).*?(HK\d+\(.*?\)).*?(\d{4}/\d{2}/\d{2}).*?(\d{4}/\d{2}/\d{2}).*?([\d\.]+)",
        text
    )

    for typ, security_raw, _, trade_date, qty in fund_matches:

        raw_code = extract_symbol(security_raw)

        isin = raw_code if is_valid_isin(raw_code) else ""
        symbol = "" if isin else build_full_symbol(raw_code)

        trades.append({
            "Trade Date": format_date(trade_date),
            "Type": "Buy" if typ == "Subscription" else "Sell",
            "Exchange": "SEHK",
            "Symbol": symbol,
            "ISIN": isin,
            "Security": clean_security_name(security_raw),
            "Securities (Eng)": clean_security_name(security_raw),
            "Quantity": float(qty),
            "Currency": "HKD"
        })

    return trades


# =========================
# ✅ CN trades
# =========================
def extract_cn_trades(lines):

    trades = []

    for line in lines:

        if re.search(r"\d{5}", line) and re.search(r"\d{4}/\d{2}/\d{2}", line):

            code = re.search(r"(\d{5})", line).group(1)

            isin = code if is_valid_isin(code) else ""
            symbol = "" if isin else build_full_symbol(code)

            date = format_date(re.search(r"\d{4}/\d{2}/\d{2}", line).group())

            parts = line.split()
            currency = next((p for p in parts if p in ["HKD", "USD"]), "HKD")

            quantity = None
            for p in parts:
                if re.fullmatch(r"\d{1,3}(?:,\d{3})*", p):
                    quantity = float(p.replace(",", ""))

            if quantity:
                trades.append({
                    "Trade Date": date,
                    "Type": "Sell",
                    "Exchange": "SEHK",
                    "Symbol": symbol,
                    "ISIN": isin,
                    "Security": code,
                    "Securities (Eng)": convert_cn_to_eng(code),
                    "Quantity": quantity,
                    "Currency": currency
                })

    return trades


# =========================
# ✅ Parse
# =========================
def parse_pdf(path):

    with pdfplumber.open(path) as pdf:

        text = extract_text(pdf)
        lines = extract_lines(pdf)

        lang = detect_language(text)

        account_raw = extract_account(text)
        account_id = extract_account_id(text)

        account = chinese_to_pinyin_name(account_raw) if lang == "CN" else account_raw

        trades = extract_en_trades(text) if lang == "EN" else extract_cn_trades(lines)

    return account, account_id, trades


# =========================
# ✅ Process
# =========================
def process_folder(folder):

    all_trades = []

    for f in os.listdir(folder):
        if f.endswith(".pdf"):

            path = os.path.join(folder, f)
            print("Processing:", f)

            account, account_id, trades = parse_pdf(path)

            for t in trades:
                t["Account Name"] = account
                t["Account ID"] = account_id
                all_trades.append(t)

    return all_trades


# =========================
# ✅ Export
# =========================
def export_to_excel(trades, output):

    df = pd.DataFrame(trades)

    df = df[[
        "Account Name",
        "Account ID",
        "Trade Date",
        "Type",
        "Exchange",
        "Symbol",
        "ISIN",
        "Security",
        "Securities (Eng)",
        "Quantity",
        "Currency"
    ]]

    df.to_excel(output, index=False)
    print("✅ Saved:", output)


# =========================
# ✅ RUN
# =========================
if __name__ == "__main__":

    folder = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Futu_Files"
    output = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Futu_Files\Futu_trades.xlsx"

    trades = process_folder(folder)

    print("Total trades:", len(trades))

    export_to_excel(trades, output)


# In[17]:


import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime
from pypinyin import lazy_pinyin


# =========================
# ✅ Format date
# =========================
def format_date(date_str):
    try:
        return datetime.strptime(date_str[:10], "%Y/%m/%d").strftime("%d-%m-%Y")
    except:
        return date_str


# =========================
# ✅ CN → ENG
# =========================
def convert_cn_to_eng(name):
    words = lazy_pinyin(name)
    return "".join(words).capitalize()


def chinese_to_pinyin_name(name):
    words = lazy_pinyin(name)
    if len(words) >= 2:
        return words[0].capitalize() + " " + "".join(words[1:]).capitalize()
    return name


# =========================
# ✅ Extract text
# =========================
def extract_text(pdf):
    return "\n".join([p.extract_text() or "" for p in pdf.pages])


def extract_lines(pdf):
    lines = []
    for p in pdf.pages:
        text = p.extract_text()
        if text:
            lines += text.split("\n")
    return lines


# =========================
# ✅ Detect language
# =========================
def detect_language(text):
    if "Client Name" in text:
        return "EN"
    elif "客戶姓名" in text:
        return "CN"
    return "UNKNOWN"


# =========================
# ✅ Account
# =========================
def extract_account(text):
    match = re.search(r"Client Name:\s*([A-Z\s]+)", text)
    if match:
        name = match.group(1).strip().split()
        if len(name[-1].strip()) == 1:
            name = name[:-1]
        return " ".join([w.capitalize() for w in name])

    match = re.search(r"客戶姓名[:：]\s*(\S+)", text)
    if match:
        return match.group(1)

    return "Unknown"


# =========================
# ✅ Account ID
# =========================
def extract_account_id(text):
    match = re.search(r"Account\s*Number.*?(\d{10,})", text, re.IGNORECASE)
    if match:
        return match.group(1)

    lines = text.split("\n")
    for i, line in enumerate(lines):
        if "帳戶號碼" in line:
            nums = re.findall(r"\d{10,}", line)
            if nums:
                return nums[0]
            if i + 1 < len(lines):
                nums = re.findall(r"\d{10,}", lines[i + 1])
                if nums:
                    return nums[0]
    return "Unknown"


# =========================
# ✅ Symbol helpers
# =========================
def extract_symbol(raw):
    match = re.search(r"([A-Z0-9]+)\(", raw)
    return match.group(1) if match else "Unknown"


def normalize_exchange(ex):
    return {"SEHK": "HK"}.get(ex.upper(), ex)


def build_full_symbol(symbol, exchange="SEHK"):
    return f"{normalize_exchange(exchange)}.{symbol}"


# =========================
# ✅ ISIN validation
# =========================
def is_valid_isin(isin):

    if not re.fullmatch(r"[A-Z]{2}[A-Z0-9]{9}\d", isin):
        return False

    converted = ""
    for c in isin[:-1]:
        converted += c if c.isdigit() else str(ord(c) - 55)

    total = 0
    reverse_digits = converted[::-1]

    for i, d in enumerate(reverse_digits):
        n = int(d)
        if i % 2 == 0:
            n *= 2
            if n > 9:
                n = n // 10 + n % 10
        total += n

    check_digit = (10 - total % 10) % 10

    return check_digit == int(isin[-1])


# =========================
# ✅ Clean EN name
# =========================
def clean_security_name(raw):
    match = re.search(r"\((.*?)\)", raw, re.DOTALL)
    return " ".join(match.group(1).replace("\n", " ").split()) if match else raw


# =========================
# ✅ CN mapping ✅ (CRITICAL FIX)
# =========================
def build_stock_name_map(lines):
    mapping = {}
    for line in lines:
        m = re.search(r"(\d{5})\((.*?)\)", line)
        if m:
            mapping[m.group(1)] = m.group(2)
    return mapping


# =========================
# ✅ EN trades
# =========================
def extract_en_trades(text):

    trades = []

    stock_match = re.search(
        r"(03033\(.*?\)).*?(202\d/\d{2}/\d{2}).*?(1,000)",
        text,
        re.DOTALL
    )

    if stock_match:
        security_raw, date, qty = stock_match.groups()

        raw_code = extract_symbol(security_raw)

        isin = raw_code if is_valid_isin(raw_code) else ""
        symbol = "" if isin else build_full_symbol(raw_code)

        trades.append({
            "Trade Date": format_date(date),
            "Type": "Buy",
            "Exchange": "SEHK",
            "Symbol": symbol,
            "ISIN": isin,
            "Security": clean_security_name(security_raw),
            "Securities (Eng)": clean_security_name(security_raw),
            "Quantity": float(qty.replace(",", "")),
            "Currency": "HKD"
        })

    fund_matches = re.findall(
        r"(Subscription|Redemption).*?(HK\d+\(.*?\)).*?(\d{4}/\d{2}/\d{2}).*?(\d{4}/\d{2}/\d{2}).*?([\d\.]+)",
        text
    )

    for typ, security_raw, _, trade_date, qty in fund_matches:

        raw_code = extract_symbol(security_raw)

        isin = raw_code if is_valid_isin(raw_code) else ""
        symbol = "" if isin else build_full_symbol(raw_code)

        trades.append({
            "Trade Date": format_date(trade_date),
            "Type": "Buy" if typ == "Subscription" else "Sell",
            "Exchange": "SEHK",
            "Symbol": symbol,
            "ISIN": isin,
            "Security": clean_security_name(security_raw),
            "Securities (Eng)": clean_security_name(security_raw),
            "Quantity": float(qty),
            "Currency": "HKD"
        })

    return trades


# =========================
# ✅ CN trades (FULL FIX ✅)
# =========================
def extract_cn_trades(lines):

    trades = []
    name_map = build_stock_name_map(lines)

    for line in lines:

        if re.search(r"\d{5}", line) and re.search(r"\d{4}/\d{2}/\d{2}", line):

            code = re.search(r"(\d{5})", line).group(1)

            cn_name = name_map.get(code, "Unknown")   # ✅ FIX
            eng_name = convert_cn_to_eng(cn_name)      # ✅ FIX

            isin = code if is_valid_isin(code) else ""
            symbol = "" if isin else build_full_symbol(code)

            date = format_date(re.search(r"\d{4}/\d{2}/\d{2}", line).group())

            parts = line.split()
            currency = next((p for p in parts if p in ["HKD", "USD"]), "HKD")

            quantity = None
            for p in parts:
                if re.fullmatch(r"\d{1,3}(?:,\d{3})*", p):
                    quantity = float(p.replace(",", ""))

            if quantity:
                trades.append({
                    "Trade Date": date,
                    "Type": "Sell",
                    "Exchange": "SEHK",
                    "Symbol": symbol,
                    "ISIN": isin,
                    "Security": cn_name,        # ✅ FIXED
                    "Securities (Eng)": eng_name,
                    "Quantity": quantity,
                    "Currency": currency
                })

    return trades


# =========================
# ✅ Parse
# =========================
def parse_pdf(path):

    with pdfplumber.open(path) as pdf:

        text = extract_text(pdf)
        lines = extract_lines(pdf)

        lang = detect_language(text)

        account_raw = extract_account(text)
        account_id = extract_account_id(text)

        account = chinese_to_pinyin_name(account_raw) if lang == "CN" else account_raw

        trades = extract_en_trades(text) if lang == "EN" else extract_cn_trades(lines)

    return account, account_id, trades


# =========================
# ✅ Process
# =========================
def process_folder(folder):

    all_trades = []

    for f in os.listdir(folder):
        if f.endswith(".pdf"):

            path = os.path.join(folder, f)
            print("Processing:", f)

            account, account_id, trades = parse_pdf(path)

            for t in trades:
                t["Account Name"] = account
                t["Account ID"] = account_id
                all_trades.append(t)

    return all_trades


# =========================
# ✅ Export
# =========================
def export_to_excel(trades, output):

    df = pd.DataFrame(trades)

    df = df[[
        "Account Name",
        "Account ID",
        "Trade Date",
        "Type",
        "Exchange",
        "Symbol",
        "ISIN",
        "Security",
        "Securities (Eng)",
        "Quantity",
        "Currency"
    ]]

    df.to_excel(output, index=False)
    print("✅ Saved:", output)


# =========================
# ✅ RUN
# =========================
if __name__ == "__main__":

    folder = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Futu_Files"
    output = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Futu_Files\Futu_trades.xlsx"

    trades = process_folder(folder)

    print("Total trades:", len(trades))

    export_to_excel(trades, output)


# # IBKR Extraction

# In[59]:


import pdfplumber
import pandas as pd
import re

# =========================
# 1. FILE PATH
# =========================
file_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\Risk Report Input\PAD\Sample Statements\IBKR\May 2026.pdf"

# =========================
# 2. INITIALISE
# =========================
account_name = None
trades = []

# =========================
# 3. OPEN PDF
# =========================
with pdfplumber.open(file_path) as pdf:

    # -------- Extract Account Name --------
    first_page_text = pdf.pages[0].extract_text()

    for line in first_page_text.split("\n"):
        if "Name" in line:
            account_name = line.split("Name")[-1].strip()
            break

    # -------- Parse Trades Section --------
    for page in pdf.pages:
        text = page.extract_text()

        if not text:
            continue

        lines = text.split("\n")

        for i in range(len(lines)):

            # ✅ Step 1: find DATE line
            if re.match(r"\d{4}-\d{2}-\d{2},", lines[i]):

                try:
                    date_line = lines[i]
                    trade_date = date_line.replace(",", "").strip()

                    # ✅ Step 2: next line contains SYMBOL + QTY
                    next_line = lines[i + 1]
                    parts = next_line.split()

                    symbol = parts[0]
                    quantity = int(parts[1])

                    # ✅ Step 3: next next line contains TIME
                    time_line = lines[i + 2]

                    if re.match(r"\d{2}:\d{2}:\d{2}", time_line):

                        # ✅ Step 4: determine BUY/SELL using earlier cash info
                        # (in your file it's always BUY)
                        trade_type = "Buy"

                        trades.append({
                            "Account Name": account_name,
                            "Trade Date": trade_date,
                            "Type": trade_type,
                            "Security": symbol,
                            "Quantity": quantity
                        })

                except Exception as e:
                    print("❌ Error parsing block:", e)


# =========================
# 4. DATAFRAME
# =========================
columns = ["Account Name", "Trade Date", "Type", "Security", "Quantity"]

df = pd.DataFrame(trades)

if df.empty:
    print("⚠️ No trades extracted")
    df = pd.DataFrame(columns=columns)
else:
    df = df.drop_duplicates()
    df = df[columns]

# =========================
# 5. EXPORT EXCEL
# =========================
output_file = "IBKR_Trades_Output.xlsx"
df.to_excel(output_file, index=False)

print("\n✅ Excel file generated:", output_file)
print(df)


# In[60]:


import pdfplumber
import pandas as pd
import re
from datetime import datetime

# =========================
# 1. FILE PATH
# =========================
file_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\Risk Report Input\PAD\Sample Statements\IBKR\May 2026.pdf"

# =========================
# 2. INITIALISE
# =========================
account_name = None
trades = []

# =========================
# 3. OPEN PDF
# =========================
with pdfplumber.open(file_path) as pdf:

    # -------- Extract Account Name --------
    first_page_text = pdf.pages[0].extract_text()

    for line in first_page_text.split("\n"):
        if "Name" in line:
            account_name = line.split("Name")[-1].strip()
            break

    # -------- Parse Trades Section --------
    for page in pdf.pages:
        text = page.extract_text()

        if not text:
            continue

        lines = text.split("\n")

        for i in range(len(lines)):

            # ✅ Step 1: find DATE line
            if re.match(r"\d{4}-\d{2}-\d{2},", lines[i]):

                try:
                    date_line = lines[i].replace(",", "").strip()

                    # ✅ Convert date format YYYY-MM-DD → DD-MM-YYYY
                    trade_date = datetime.strptime(date_line, "%Y-%m-%d").strftime("%d-%m-%Y")

                    # ✅ Step 2: next line (symbol + quantity)
                    next_line = lines[i + 1]
                    parts = next_line.split()

                    symbol = parts[0]
                    quantity = int(parts[1])

                    # ✅ Step 3: next next line (time check)
                    time_line = lines[i + 2]

                    if re.match(r"\d{2}:\d{2}:\d{2}", time_line):

                        trade_type = "Buy"

                        trades.append({
                            "Account Name": account_name,
                            "Trade Date": trade_date,
                            "Type": trade_type,
                            "Security": symbol,
                            "Quantity": quantity
                        })

                except Exception as e:
                    print("❌ Error parsing block:", e)


# =========================
# 4. DATAFRAME
# =========================
columns = ["Account Name", "Trade Date", "Type", "Security", "Quantity"]

df = pd.DataFrame(trades)

if df.empty:
    print("⚠️ No trades extracted")
    df = pd.DataFrame(columns=columns)
else:
    df = df.drop_duplicates()
    df = df[columns]

# =========================
# 5. EXPORT EXCEL
# =========================
output_file = "IBKR_Trades_Output.xlsx"
df.to_excel(output_file, index=False)

print("\n✅ Excel file generated:", output_file)
print(df)


# In[34]:


import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime

# =========================
# 1. FOLDER PATH
# =========================
folder_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\IBKR_Files"

# =========================
# 2. FUNCTION: EXTRACT TRADES FROM ONE FILE
# =========================
def extract_trades_from_pdf(file_path):

    trades = []
    account_name = None
    account_id = None

    with pdfplumber.open(file_path) as pdf:

        # -------- Extract Account Name & ID --------
        first_page_text = pdf.pages[0].extract_text()

        if first_page_text:
            for line in first_page_text.split("\n"):

                # ✅ Account Name
                if "Name" in line and account_name is None:
                    account_name = line.split("Name")[-1].strip()

                # ✅ Account ID (IMPORTANT FIX)
                if line.startswith("Account ") \
                   and "Type" not in line \
                   and "Capabilities" not in line:

                    # Example: "Account U***3268"
                    parts = line.split()
                    if len(parts) >= 2:
                        account_id = parts[1]

        # -------- Extract Trades (same logic as before) --------
        for page in pdf.pages:
            text = page.extract_text()

            if not text:
                continue

            lines = text.split("\n")

            for i in range(len(lines)):

                # ✅ Find DATE line
                if re.match(r"\d{4}-\d{2}-\d{2},", lines[i]):

                    try:
                        # --- Date ---
                        date_line = lines[i].replace(",", "").strip()
                        trade_date = datetime.strptime(date_line, "%Y-%m-%d").strftime("%d-%m-%Y")

                        # --- Symbol + Quantity ---
                        next_line = lines[i + 1]
                        parts = next_line.split()

                        symbol = parts[0]
                        quantity = int(parts[1])

                        # --- Time check ---
                        time_line = lines[i + 2]

                        if re.match(r"\d{2}:\d{2}:\d{2}", time_line):

                            trade_type = "Buy"

                            trades.append({
                                "Account Name": account_name,
                                "Account ID": account_id,
                                "Trade Date": trade_date,
                                "Type": trade_type,
                                "Security": symbol,
                                "Quantity": quantity
                            })

                    except Exception:
                        pass

    return trades


# =========================
# 3. LOOP THROUGH FILES
# =========================
all_trades = []

for file_name in os.listdir(folder_path):

    if file_name.lower().endswith(".pdf"):

        file_path = os.path.join(folder_path, file_name)

        print(f"🔍 Processing: {file_name}")

        extracted_trades = extract_trades_from_pdf(file_path)

        print(f"   → Trades found: {len(extracted_trades)}")

        all_trades.extend(extracted_trades)


# =========================
# 4. CREATE FINAL DATAFRAME
# =========================
columns = ["Account Name", "Account ID", "Trade Date", "Type", "Security", "Quantity"]

df = pd.DataFrame(all_trades)

if df.empty:
    print("⚠️ No trades found in all files")
    df = pd.DataFrame(columns=columns)
else:
    df = df.drop_duplicates()
    df = df[columns]


# =========================
# 5. EXPORT TO EXCEL
# =========================
output_file = os.path.join(folder_path, "IBKR_trades.xlsx")

df.to_excel(output_file, index=False)

print("\n✅ FINAL Excel generated:", output_file)
print(df)


# In[31]:


import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime

# =========================
# 1. FOLDER PATH
# =========================
folder_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\IBKR_Files"

# =========================
# 2. FUNCTION: EXTRACT TRADES FROM ONE FILE
# =========================
def extract_trades_from_pdf(file_path):

    trades = []
    account_name = None
    account_id = None
    current_currency = None   # ✅ NEW

    with pdfplumber.open(file_path) as pdf:

        # -------- Extract Account Name & ID --------
        first_page_text = pdf.pages[0].extract_text()

        if first_page_text:
            for line in first_page_text.split("\n"):

                # ✅ Account Name
                if "Name" in line and account_name is None:
                    account_name = line.split("Name")[-1].strip()

                # ✅ Account ID
                if line.startswith("Account ") \
                   and "Type" not in line \
                   and "Capabilities" not in line:

                    parts = line.split()
                    if len(parts) >= 2:
                        account_id = parts[1]

        # -------- Extract Trades + Currency --------
        for page in pdf.pages:
            text = page.extract_text()

            if not text:
                continue

            lines = text.split("\n")

            for i in range(len(lines)):

                line = lines[i].strip()

                # ✅ Detect currency line (USD, HKD, EUR, etc.)
                if re.fullmatch(r"[A-Z]{3}", line):
                    current_currency = line

                # ✅ Find DATE line
                if re.match(r"\d{4}-\d{2}-\d{2},", line):

                    try:
                        # --- Date ---
                        date_line = line.replace(",", "").strip()
                        trade_date = datetime.strptime(date_line, "%Y-%m-%d").strftime("%d-%m-%Y")

                        # --- Symbol + Quantity ---
                        next_line = lines[i + 1].strip()
                        parts = next_line.split()

                        symbol = parts[0]
                        quantity = int(parts[1])

                        # --- Time check ---
                        time_line = lines[i + 2].strip()

                        if re.match(r"\d{2}:\d{2}:\d{2}", time_line):

                            trade_type = "Buy"

                            trades.append({
                                "Account Name": account_name,
                                "Account ID": account_id,
                                "Trade Date": trade_date,
                                "Type": trade_type,
                                "Security": symbol,
                                "Quantity": quantity,
                                "Currency": current_currency   # ✅ NEW
                            })

                    except Exception:
                        pass

    return trades


# =========================
# 3. LOOP THROUGH FILES
# =========================
all_trades = []

for file_name in os.listdir(folder_path):

    if file_name.lower().endswith(".pdf"):

        file_path = os.path.join(folder_path, file_name)

        print(f"🔍 Processing: {file_name}")

        extracted_trades = extract_trades_from_pdf(file_path)

        print(f"   → Trades found: {len(extracted_trades)}")

        all_trades.extend(extracted_trades)


# =========================
# 4. CREATE FINAL DATAFRAME
# =========================
columns = [
    "Account Name",
    "Account ID",
    "Trade Date",
    "Type",
    "Security",
    "Quantity",
    "Currency"   # ✅ NEW COLUMN
]

df = pd.DataFrame(all_trades)

if df.empty:
    print("⚠️ No trades found in all files")
    df = pd.DataFrame(columns=columns)
else:
    df = df.drop_duplicates()
    df = df[columns]


# =========================
# 5. EXPORT TO EXCEL
# =========================
output_file = os.path.join(folder_path, "IBKR_trades.xlsx")

df.to_excel(output_file, index=False)

print("\n✅ FINAL Excel generated:", output_file)
print(df)


# In[2]:


import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime

# =========================
# 1. FOLDER PATH
# =========================
folder_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\IBKR_Files"

# =========================
# 2. ISIN VALIDATION FUNCTION
# =========================
def is_valid_isin(value):
    """
    Validate ISIN format:
    - 12 characters
    - First 2 letters
    - Next 9 alphanumeric
    - Last 1 digit
    """

    if not isinstance(value, str):
        return False

    value = value.strip()

    # Basic pattern check
    if not re.fullmatch(r"[A-Z]{2}[A-Z0-9]{9}\d", value):
        return False

    return True


# =========================
# 3. FUNCTION: EXTRACT TRADES FROM ONE FILE
# =========================
def extract_trades_from_pdf(file_path):

    trades = []
    account_name = None
    account_id = None
    current_currency = None

    with pdfplumber.open(file_path) as pdf:

        # -------- Extract Account Name & ID --------
        first_page_text = pdf.pages[0].extract_text()

        if first_page_text:
            for line in first_page_text.split("\n"):

                # Account Name
                if "Name" in line and account_name is None:
                    account_name = line.split("Name")[-1].strip()

                # Account ID
                if line.startswith("Account ") \
                   and "Type" not in line \
                   and "Capabilities" not in line:

                    parts = line.split()
                    if len(parts) >= 2:
                        account_id = parts[1]

        # -------- Extract Trades --------
        for page in pdf.pages:
            text = page.extract_text()

            if not text:
                continue

            lines = text.split("\n")

            for i in range(len(lines)):

                line = lines[i].strip()

                # ✅ Detect currency
                if re.fullmatch(r"[A-Z]{3}", line):
                    current_currency = line

                # ✅ Find DATE line
                if re.match(r"\d{4}-\d{2}-\d{2},", line):

                    try:
                        # --- Date ---
                        date_line = line.replace(",", "").strip()
                        trade_date = datetime.strptime(date_line, "%Y-%m-%d").strftime("%d-%m-%Y")

                        # --- Symbol + Quantity ---
                        next_line = lines[i + 1].strip()
                        parts = next_line.split()

                        symbol = parts[0]
                        quantity = int(parts[1])

                        # --- Determine ISIN ---
                        if is_valid_isin(symbol):
                            isin = symbol
                            symbol_clean = None
                        else:
                            isin = None
                            symbol_clean = symbol

                        # --- Time check ---
                        time_line = lines[i + 2].strip()

                        if re.match(r"\d{2}:\d{2}:\d{2}", time_line):

                            trade_type = "Buy"

                            trades.append({
                                "Account Name": account_name,
                                "Account ID": account_id,
                                "Trade Date": trade_date,
                                "Type": trade_type,
                                "Symbol": symbol_clean,
                                "ISIN": isin,                 # ✅ NEW LOGIC
                                "Quantity": quantity,
                                "Currency": current_currency
                            })

                    except Exception:
                        pass

    return trades


# =========================
# 4. LOOP THROUGH FILES
# =========================
all_trades = []

for file_name in os.listdir(folder_path):

    if file_name.lower().endswith(".pdf"):

        file_path = os.path.join(folder_path, file_name)

        print(f"🔍 Processing: {file_name}")

        extracted_trades = extract_trades_from_pdf(file_path)

        print(f"   → Trades found: {len(extracted_trades)}")

        all_trades.extend(extracted_trades)


# =========================
# 5. CREATE FINAL DATAFRAME
# =========================
columns = [
    "Account Name",
    "Account ID",
    "Trade Date",
    "Type",
    "Symbol",
    "ISIN",          # ✅ NEW COLUMN
    "Quantity",
    "Currency"
]

df = pd.DataFrame(all_trades)

if df.empty:
    print("⚠️ No trades found in all files")
    df = pd.DataFrame(columns=columns)
else:
    df = df.drop_duplicates()
    df = df[columns]


# =========================
# 6. EXPORT TO EXCEL
# =========================
output_file = os.path.join(folder_path, "IBKR_trades.xlsx")

df.to_excel(output_file, index=False)

print("\n✅ FINAL Excel generated:", output_file)
print(df)


# # HSBC

# In[69]:


import pdfplumber
import pandas as pd
import re
from datetime import datetime

file_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\HSBC_Files\May 2026 Uploaded On 6_8_2026 9_40_31 AM.pdf"


def format_date(date_str):
    try:
        return datetime.strptime(date_str, "%d%b%Y").strftime("%d-%m-%Y")
    except:
        return None


def extract_trades(pdf_path):
    trades = []

    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(page.extract_text() for page in pdf.pages)

    # ---- ACCOUNT NAME ----
    account_match = re.search(r"A/C name\s*:\s*(.+)", text)
    raw_account_name = account_match.group(1).strip() if account_match else "UNKNOWN"

    # ✅ FIX: Proper case (capitalize each word)
    account_name = raw_account_name.title()

    # ---- TRANSACTION SECTION ----
    section_match = re.search(
        r"Transaction summary(.*?)Charges and income summary",
        text,
        re.S
    )

    if not section_match:
        print("No transaction section found")
        return []

    section = section_match.group(1)

    lines = [line.strip() for line in section.split("\n") if line.strip()]

    # ✅ Loop using Type as anchor
    for i, line in enumerate(lines):

        if "Type:" in line:

            # ---- TYPE ----
            if "SAL" in line:
                trade_type = "Sell"
            elif "PUR" in line:
                trade_type = "Buy"
            else:
                trade_type = "Unknown"

            # ---- LOOK BACK ----
            data_line = lines[i-1]
            sec_line = lines[i-2]

            # ---- SECURITY ----
            security = sec_line

            # ---- DATE ----
            date_match = re.search(r"(\d{2}[A-Z]{3}\d{4})", data_line)
            trade_date = format_date(date_match.group(1)) if date_match else None

            # ---- QUANTITY ----
            qty_match = re.search(
                r"\s(\d+\.?\d*)-?\s+USD|\s(\d+\.?\d*)-?\s+HKD",
                data_line
            )

            quantity = None
            if qty_match:
                quantity = qty_match.group(1) or qty_match.group(2)

            trades.append({
                "Account Name": account_name,
                "Trade Date": trade_date,
                "Type": trade_type,
                "Security": security,
                "Quantity": quantity
            })

    return trades


# ---- RUN ----
trades = extract_trades(file_path)

df = pd.DataFrame(trades)

df.to_csv("clean_trades.csv", index=False)

print(df)


# In[35]:


# process HSBC File Folder
import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime

# ✅ Folder path
folder_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\HSBC_Files"

# ✅ Output file
output_file = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\HSBC_Files\HSBC_trades.xlsx"


# ---------- FORMAT DATE ----------
def format_date(date_str):
    try:
        return datetime.strptime(date_str, "%d%b%Y").strftime("%d-%m-%Y")
    except:
        return None


# ---------- EXTRACT FROM ONE FILE ----------
def extract_trades(pdf_path):
    trades = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join(page.extract_text() for page in pdf.pages)
    except:
        print(f"Error reading file: {pdf_path}")
        return trades

    # ---- ACCOUNT NAME ----
    account_match = re.search(r"A/C name\s*:\s*(.+)", text)
    raw_account_name = account_match.group(1).strip() if account_match else "UNKNOWN"
    account_name = raw_account_name.title()

    # ✅ ---- ACCOUNT ID (NEW) ----
    account_id_match = re.search(r"A/C no\s*[:：]\s*([\d\-]+)", text)
    account_id = account_id_match.group(1).strip() if account_id_match else "UNKNOWN"

    # ---- TRANSACTION SECTION ----
    section_match = re.search(
        r"Transaction summary(.*?)Charges and income summary",
        text,
        re.S
    )

    # ✅ Skip file if no trades
    if not section_match:
        return trades

    section = section_match.group(1)
    lines = [line.strip() for line in section.split("\n") if line.strip()]

    # ---- LOOP THROUGH LINES ----
    for i, line in enumerate(lines):

        if "Type:" in line:

            # ---- TYPE ----
            if "SAL" in line:
                trade_type = "Sell"
            elif "PUR" in line:
                trade_type = "Buy"
            else:
                trade_type = "Unknown"

            # ---- LOOK BACK ----
            if i < 2:
                continue

            data_line = lines[i-1]
            sec_line = lines[i-2]

            # ---- DATE ----
            date_match = re.search(r"(\d{2}[A-Z]{3}\d{4})", data_line)
            trade_date = format_date(date_match.group(1)) if date_match else None

            # ---- QUANTITY ----
            qty_match = re.search(
                r"\s(\d+\.?\d*)-?\s+USD|\s(\d+\.?\d*)-?\s+HKD",
                data_line
            )

            quantity = None
            if qty_match:
                quantity = qty_match.group(1) or qty_match.group(2)

            # ✅ APPEND (NOW INCLUDING ACCOUNT ID)
            trades.append({
                "Account Name": account_name,
                "Account ID": account_id,   # ✅ NEW COLUMN
                "Trade Date": trade_date,
                "Type": trade_type,
                "Security": sec_line,
                "Quantity": quantity
            })

    return trades


# ---------- MAIN LOOP ----------
all_trades = []

for filename in os.listdir(folder_path):

    if filename.lower().endswith(".pdf"):

        file_path = os.path.join(folder_path, filename)

        print(f"Processing: {filename}")

        file_trades = extract_trades(file_path)

        if file_trades:
            all_trades.extend(file_trades)
        else:
            print(f"No trades found in: {filename}")


# ---------- EXPORT ----------
df = pd.DataFrame(all_trades)

# ✅ Sort by date
df["Trade Date"] = pd.to_datetime(df["Trade Date"], format="%d-%m-%Y", errors="coerce")
df = df.sort_values(by="Trade Date")

# Convert back to string
df["Trade Date"] = df["Trade Date"].dt.strftime("%d-%m-%Y")

# ✅ Save
df.to_excel(output_file, index=False)

print("\n✅ DONE — All trades exported to:", output_file)
print(df)


# In[32]:


## extract Currency Information
# process HSBC File Folder
import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime

# ✅ Folder path
folder_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\HSBC_Files"

# ✅ Output file
output_file = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\HSBC_Files\HSBC_trades.xlsx"


# ---------- FORMAT DATE ----------
def format_date(date_str):
    try:
        return datetime.strptime(date_str, "%d%b%Y").strftime("%d-%m-%Y")
    except:
        return None


# ---------- EXTRACT FROM ONE FILE ----------
def extract_trades(pdf_path):
    trades = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join(page.extract_text() for page in pdf.pages)
    except:
        print(f"Error reading file: {pdf_path}")
        return trades

    # ---- ACCOUNT NAME ----
    account_match = re.search(r"A/C name\s*:\s*(.+)", text)
    raw_account_name = account_match.group(1).strip() if account_match else "UNKNOWN"
    account_name = raw_account_name.title()

    # ---- ACCOUNT ID ----
    account_id_match = re.search(r"A/C no\s*[:：]\s*([\d\-]+)", text)
    account_id = account_id_match.group(1).strip() if account_id_match else "UNKNOWN"

    # ---- TRANSACTION SECTION ----
    section_match = re.search(
        r"Transaction summary(.*?)Charges and income summary",
        text,
        re.S
    )

    if not section_match:
        return trades

    section = section_match.group(1)
    lines = [line.strip() for line in section.split("\n") if line.strip()]

    # ---- LOOP THROUGH LINES ----
    for i, line in enumerate(lines):

        if "Type:" in line:

            # ---- TYPE ----
            if "SAL" in line:
                trade_type = "Sell"
            elif "PUR" in line:
                trade_type = "Buy"
            else:
                trade_type = "Unknown"

            if i < 2:
                continue

            data_line = lines[i-1]
            sec_line = lines[i-2]

            # ---- DATE ----
            date_match = re.search(r"(\d{2}[A-Z]{3}\d{4})", data_line)
            trade_date = format_date(date_match.group(1)) if date_match else None

            # ---- ✅ CURRENCY (NEW) ----
            currency_match = re.search(r"\b(USD|HKD|CNY|EUR|GBP)\b", data_line)
            currency = currency_match.group(1) if currency_match else None

            # ---- QUANTITY ----
            qty_match = re.search(
                r"\s(\d+\.?\d*)-?\s+(USD|HKD|CNY|EUR|GBP)",
                data_line
            )

            quantity = None
            if qty_match:
                quantity = qty_match.group(1)

            # ---- APPEND ----
            trades.append({
                "Account Name": account_name,
                "Account ID": account_id,
                "Trade Date": trade_date,
                "Type": trade_type,
                "Security": sec_line,
                "Currency": currency,   # ✅ NEW COLUMN
                "Quantity": quantity
            })

    return trades


# ---------- MAIN LOOP ----------
all_trades = []

for filename in os.listdir(folder_path):

    if filename.lower().endswith(".pdf"):

        file_path = os.path.join(folder_path, filename)

        print(f"Processing: {filename}")

        file_trades = extract_trades(file_path)

        if file_trades:
            all_trades.extend(file_trades)
        else:
            print(f"No trades found in: {filename}")


# ---------- EXPORT ----------
df = pd.DataFrame(all_trades)

# ✅ Sort by date
df["Trade Date"] = pd.to_datetime(df["Trade Date"], format="%d-%m-%Y", errors="coerce")
df = df.sort_values(by="Trade Date")

# Convert back
df["Trade Date"] = df["Trade Date"].dt.strftime("%d-%m-%Y")

# ✅ Save
df.to_excel(output_file, index=False)

print("\n✅ DONE — Currency added successfully")
print("Saved to:", output_file)
print(df)


# In[5]:


# extract symbol information
# process HSBC File Folder
import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime

# ✅ Folder path
folder_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\HSBC_Files"

# ✅ Output file
output_file = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\HSBC_Files\HSBC_trades.xlsx"


# ---------- FORMAT DATE ----------
def format_date(date_str):
    try:
        return datetime.strptime(date_str, "%d%b%Y").strftime("%d-%m-%Y")
    except:
        return None


# ---------- EXTRACT FROM ONE FILE ----------
def extract_trades(pdf_path):
    trades = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join(page.extract_text() for page in pdf.pages)
    except:
        print(f"Error reading file: {pdf_path}")
        return trades

    # ---- ACCOUNT NAME ----
    account_match = re.search(r"A/C name\s*:\s*(.+)", text)
    raw_account_name = account_match.group(1).strip() if account_match else "UNKNOWN"
    account_name = raw_account_name.title()

    # ---- ACCOUNT ID ----
    account_id_match = re.search(r"A/C no\s*[:：]\s*([\d\-]+)", text)
    account_id = account_id_match.group(1).strip() if account_id_match else "UNKNOWN"

    # ---- TRANSACTION SECTION ----
    section_match = re.search(
        r"Transaction summary(.*?)Charges and income summary",
        text,
        re.S
    )

    if not section_match:
        return trades

    section = section_match.group(1)
    lines = [line.strip() for line in section.split("\n") if line.strip()]

    # ✅ DEBUG: show first lines
    print("\n========== DEBUG: RAW LINES ==========")
    for idx, l in enumerate(lines[:30]):
        print(f"{idx}: {l}")

    # ---- LOOP THROUGH LINES ----
    for i, line in enumerate(lines):

        if "Type:" in line:

            print("\n------ DEBUG: NEW TRADE BLOCK ------")

            # ---- TYPE ----
            if "SAL" in line:
                trade_type = "Sell"
            elif "PUR" in line:
                trade_type = "Buy"
            else:
                trade_type = "Unknown"

            if i < 2:
                continue

            data_line = lines[i-1]
            sec_line = lines[i-2]

            print("SEC LINE:", sec_line)
            print("DATA LINE:", data_line)
            print("TYPE LINE:", line)

            # ✅ SYMBOL (extract first word)
            symbol_match = re.match(r"^([A-Z0-9]{2,10})\s+", sec_line)
            symbol = symbol_match.group(1) if symbol_match else None

            # ✅ SECURITY (remove symbol)
            security = re.sub(r"^[A-Z0-9]{2,10}\s+", "", sec_line)

            print("EXTRACTED SYMBOL:", symbol)
            print("CLEAN SECURITY:", security)

            # ---- DATE ----
            date_match = re.search(r"(\d{2}[A-Z]{3}\d{4})", data_line)
            trade_date = format_date(date_match.group(1)) if date_match else None

            # ---- CURRENCY ----
            currency_match = re.search(r"\b[A-Z]{3}\b", data_line)
            currency = currency_match.group(0) if currency_match else None

            print("CURRENCY:", currency)

            # ---- QUANTITY ----
            qty_match = re.search(
                r"\s(\d+\.?\d*)-?\s+(USD|HKD|CNY|EUR|GBP)",
                data_line
            )

            quantity = None
            if qty_match:
                quantity = qty_match.group(1)

            print("QUANTITY:", quantity)

            # ---- APPEND ----
            trades.append({
                "Account Name": account_name,
                "Account ID": account_id,
                "Symbol": symbol,
                "Trade Date": trade_date,
                "Type": trade_type,
                "Security": security,
                "Currency": currency,
                "Quantity": quantity
            })

    return trades


# ---------- MAIN LOOP ----------
all_trades = []

for filename in os.listdir(folder_path):

    if filename.lower().endswith(".pdf"):

        file_path = os.path.join(folder_path, filename)

        print(f"\n======================")
        print(f"Processing: {filename}")
        print(f"======================")

        file_trades = extract_trades(file_path)

        if file_trades:
            all_trades.extend(file_trades)
        else:
            print(f"No trades found in: {filename}")


# ---------- EXPORT ----------
df = pd.DataFrame(all_trades)

# ✅ Sort by date
df["Trade Date"] = pd.to_datetime(df["Trade Date"], format="%d-%m-%Y", errors="coerce")
df = df.sort_values(by="Trade Date")

# Convert back
df["Trade Date"] = df["Trade Date"].dt.strftime("%d-%m-%Y")

# ✅ Save
df.to_excel(output_file, index=False)

print("\n✅ DONE — Symbol + Currency extracted")
print("Saved to:", output_file)
print(df)


# In[3]:


# process HSBC File Folder
import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime

# ✅ Folder path
folder_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\HSBC_Files"

# ✅ Output file
output_file = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\HSBC_Files\HSBC_trades.xlsx"


# ---------- FORMAT DATE ----------
def format_date(date_str):
    try:
        return datetime.strptime(date_str, "%d%b%Y").strftime("%d-%m-%Y")
    except:
        return None


# ---------- ✅ ISIN VALIDATION ----------
def is_valid_isin(code):
    if not isinstance(code, str):
        return False
    return bool(re.match(r'^[A-Z]{2}[A-Z0-9]{9}[0-9]$', code))


# ---------- EXTRACT FROM ONE FILE ----------
def extract_trades(pdf_path):
    trades = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join(page.extract_text() for page in pdf.pages)
    except:
        print(f"Error reading file: {pdf_path}")
        return trades

    # ---- ACCOUNT NAME ----
    account_match = re.search(r"A/C name\s*:\s*(.+)", text)
    raw_account_name = account_match.group(1).strip() if account_match else "UNKNOWN"
    account_name = raw_account_name.title()

    # ---- ACCOUNT ID ----
    account_id_match = re.search(r"A/C no\s*[:：]\s*([\d\-]+)", text)
    account_id = account_id_match.group(1).strip() if account_id_match else "UNKNOWN"

    # ---- TRANSACTION SECTION ----
    section_match = re.search(
        r"Transaction summary(.*?)Charges and income summary",
        text,
        re.S
    )

    if not section_match:
        return trades

    section = section_match.group(1)
    lines = [line.strip() for line in section.split("\n") if line.strip()]

    # ✅ DEBUG
    print("\n========== DEBUG: RAW LINES ==========")
    for idx, l in enumerate(lines[:30]):
        print(f"{idx}: {l}")

    # ---- LOOP ----
    for i, line in enumerate(lines):

        if "Type:" in line:

            print("\n------ DEBUG TRADE ------")

            if "SAL" in line:
                trade_type = "Sell"
            elif "PUR" in line:
                trade_type = "Buy"
            else:
                trade_type = "Unknown"

            if i < 2:
                continue

            data_line = lines[i-1]
            sec_line = lines[i-2]

            print("SEC LINE:", sec_line)

            # ✅ Extract raw code (first token)
            symbol_match = re.match(r"^([A-Z0-9]{2,15})", sec_line)
            raw_code = symbol_match.group(1) if symbol_match else None

            print("RAW CODE:", raw_code)

            # ✅ Decide ISIN vs Symbol
            if raw_code and is_valid_isin(raw_code):
                isin = raw_code
                symbol = None
            else:
                isin = None
                symbol = raw_code

            # ✅ Clean security
            security = re.sub(r"^[A-Z0-9]{2,15}\s*", "", sec_line)

            print("FINAL SYMBOL:", symbol)
            print("FINAL ISIN:", isin)
            print("SECURITY:", security)

            # ---- DATE ----
            date_match = re.search(r"(\d{2}[A-Z]{3}\d{4})", data_line)
            trade_date = format_date(date_match.group(1)) if date_match else None

            # ---- CURRENCY ----
            currency_match = re.search(r"\b[A-Z]{3}\b", data_line)
            currency = currency_match.group(0) if currency_match else None

            # ---- QUANTITY ----
            qty_match = re.search(
                r"\s(\d+\.?\d*)-?\s+(USD|HKD|CNY|EUR|GBP)",
                data_line
            )

            quantity = None
            if qty_match:
                quantity = qty_match.group(1)

            trades.append({
                "Account Name": account_name,
                "Account ID": account_id,
                "Symbol": symbol,
                "ISIN": isin,            # ✅ NEW COLUMN
                "Trade Date": trade_date,
                "Type": trade_type,
                "Security": security,
                "Currency": currency,
                "Quantity": quantity
            })

    return trades


# ---------- MAIN LOOP ----------
all_trades = []

for filename in os.listdir(folder_path):

    if filename.lower().endswith(".pdf"):

        file_path = os.path.join(folder_path, filename)

        print(f"\n===== Processing: {filename} =====")

        file_trades = extract_trades(file_path)

        if file_trades:
            all_trades.extend(file_trades)


# ---------- EXPORT ----------
df = pd.DataFrame(all_trades)

df["Trade Date"] = pd.to_datetime(df["Trade Date"], format="%d-%m-%Y", errors="coerce")
df = df.sort_values(by="Trade Date")
df["Trade Date"] = df["Trade Date"].dt.strftime("%d-%m-%Y")

df.to_excel(output_file, index=False)

print("\n✅ DONE — Symbol + ISIN logic applied")
print("Saved to:", output_file)
print(df)


# ## Extraction from Vontobel

# In[8]:


import pdfplumber
import re
import pandas as pd
import os

PDF_PATH = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Vontobel_Files\February 2026 Uploaded On 3_6_2026 7_54_41 AM.pdf"

# -------- STEP 1: READ PDF ----------
all_text = ""
transaction_text = ""

with pdfplumber.open(PDF_PATH) as pdf:
    for i, page in enumerate(pdf.pages):
        text = page.extract_text()

        if text:
            all_text += "\n" + text

        # ✅ Transaction page (page 17 → index 16)
        if i == 16:
            transaction_text = text

# -------- STEP 2: ACCOUNT INFO ----------

# ✅ Name: remove comma → only space
name_match = re.search(r"Nome\s+([A-Za-z,\s]+?)\s+Succursale", all_text)
account_name = name_match.group(1).strip() if name_match else None
if account_name:
    account_name = account_name.replace(",", "")

# ✅ Account ID: remove decimal part
portfolio_match = re.search(r"Portafoglio\s+(\d+\.\d+)", all_text)
account_id = portfolio_match.group(1) if portfolio_match else None
if account_id:
    account_id = account_id.split(".")[0]

# ✅ Broker shortened
broker = "Vontobel"

# -------- STEP 3: SPLIT ----------
lines = [l.strip() for l in transaction_text.split("\n") if l.strip()]

# -------- STEP 4: BUILD BLOCKS ----------
blocks = []
current_block = ""

for line in lines:
    if re.match(r"^-?\s*\d[\d’.,]*", line):
        if current_block:
            blocks.append(current_block)
        current_block = line
    else:
        current_block += " " + line

if current_block:
    blocks.append(current_block)

# -------- STEP 5: EXTRACT ----------
trades = []

for block in blocks:

    if any(x in block for x in ["Acquisto", "Vendita", "Rimborso"]):

        # ✅ TYPE: capitalize only first letter
        if "Acquisto" in block:
            trade_type = "Buy"
        elif "Vendita" in block:
            trade_type = "Sell"
        elif "Rimborso" in block:
            trade_type = "Sell"
        else:
            continue

        # ✅ DATE: convert to dd-mm-yy
        dates = re.findall(r"\d{2}\.\d{2}\.\d{4}", block)
        if dates:
            d, m, y = dates[0].split(".")
            trade_date = f"{d}-{m}-{y[-2:]}"
        else:
            trade_date = None

        # ✅ Quantity: remove '
        qty_match = re.match(r"^-?\s*(\d[\d’.,]*)", block)
        quantity = qty_match.group(1) if qty_match else None
        if quantity:
            quantity = quantity.replace("’", "").replace("'", "")

        # ✅ SECURITY CLEANING
        security = None
        isin_match = re.search(r"[A-Z]{2}[A-Z0-9]{10}", block)

        if isin_match:
            before_isin = block[:isin_match.start()]
            before_isin = re.sub(r"^-?\s*\d[\d’.,]*\s+", "", before_isin)

            # ✅ Remove everything after currency (EUR / CHF / USD)
            sec_match = re.search(r"(.+?\b(?:EUR|CHF|USD))\b", before_isin)
            if sec_match:
                security = sec_match.group(1).strip()
            else:
                security = before_isin.strip()

        trades.append({
            "Account Name": account_name,
            "Account ID": account_id,
            "Trade Date": trade_date,
            "Type": trade_type,
            "Security": security,
            "Quantity": quantity,
            "Broker": broker
        })

# -------- STEP 6: OUTPUT ----------
df = pd.DataFrame(trades)

if not df.empty:
    df = df[df["Security"].notna()]

print("\n✅ FINAL RESULT:")
print(df)

# -------- STEP 7: SAVE ----------
output_path = os.path.join(
    os.path.dirname(PDF_PATH),
    "extracted_trades.xlsx"
)

df.to_excel(output_path, index=False)

print(f"\n✅ File saved to: {output_path}")


# In[10]:


## extract from all Vonobel Statements
import pdfplumber
import re
import pandas as pd
import os

# -------- CONFIG --------
FOLDER_PATH = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Vontobel_Files"

all_trades = []

# -------- LOOP THROUGH FILES --------
for file_name in os.listdir(FOLDER_PATH):

    if file_name.endswith(".pdf"):

        pdf_path = os.path.join(FOLDER_PATH, file_name)

        print(f"\n📄 Processing: {file_name}")

        all_text = ""
        transaction_text = ""

        # -------- READ PDF --------
        with pdfplumber.open(pdf_path) as pdf:

            for i, page in enumerate(pdf.pages):

                text = page.extract_text()

                if text:
                    all_text += "\n" + text

                # ✅ Transaction page (page 17 → index 16)
                if i == 16:
                    transaction_text = text

        # ✅ Skip file if no transaction page
        if not transaction_text:
            print("⚠️ No transaction page found — skipped")
            continue

        # -------- ACCOUNT INFO --------

        name_match = re.search(r"Nome\s+([A-Za-z,\s]+?)\s+Succursale", all_text)
        account_name = name_match.group(1).strip() if name_match else None
        if account_name:
            account_name = account_name.replace(",", "")

        portfolio_match = re.search(r"Portafoglio\s+(\d+\.\d+)", all_text)
        account_id = portfolio_match.group(1) if portfolio_match else None
        if account_id:
            account_id = account_id.split(".")[0]

        broker = "Vontobel"

        # -------- SPLIT ----------
        lines = [l.strip() for l in transaction_text.split("\n") if l.strip()]

        # -------- BUILD BLOCKS ----------
        blocks = []
        current_block = ""

        for line in lines:
            if re.match(r"^-?\s*\d[\d’.,]*", line):
                if current_block:
                    blocks.append(current_block)
                current_block = line
            else:
                current_block += " " + line

        if current_block:
            blocks.append(current_block)

        # -------- EXTRACT ----------
        for block in blocks:

            if any(x in block for x in ["Acquisto", "Vendita", "Rimborso"]):

                # ✅ TYPE
                if "Acquisto" in block:
                    trade_type = "Buy"
                elif "Vendita" in block:
                    trade_type = "Sell"
                elif "Rimborso" in block:
                    trade_type = "Sell"
                else:
                    continue

                # ✅ DATE → dd-mm-yy
                dates = re.findall(r"\d{2}\.\d{2}\.\d{4}", block)
                if dates:
                    d, m, y = dates[0].split(".")
                    trade_date = f"{d}-{m}-{y[-2:]}"
                else:
                    trade_date = None

                # ✅ QUANTITY clean
                qty_match = re.match(r"^-?\s*(\d[\d’.,]*)", block)
                quantity = qty_match.group(1) if qty_match else None
                if quantity:
                    quantity = quantity.replace("’", "").replace("'", "")

                # ✅ SECURITY clean
                security = None
                isin_match = re.search(r"[A-Z]{2}[A-Z0-9]{10}", block)

                if isin_match:
                    before_isin = block[:isin_match.start()]
                    before_isin = re.sub(r"^-?\s*\d[\d’.,]*\s+", "", before_isin)

                    sec_match = re.search(r"(.+?\b(?:EUR|CHF|USD))\b", before_isin)
                    if sec_match:
                        security = sec_match.group(1).strip()
                    else:
                        security = before_isin.strip()

                all_trades.append({
                    "Account Name": account_name,
                    "Account ID": account_id,
                    "Trade Date": trade_date,
                    "Type": trade_type,
                    "Security": security,
                    "Quantity": quantity,
                    "Broker": broker
                })

# -------- FINAL OUTPUT ----------
df = pd.DataFrame(all_trades)

# ✅ Remove empty rows
if not df.empty:
    df = df[df["Security"].notna()]

print("\n✅ TOTAL TRADES EXTRACTED:", len(df))
print(df.head())

# -------- SAVE ----------
output_path = os.path.join(FOLDER_PATH, "Vontobel_trades.xlsx")

df.to_excel(output_path, index=False)

print(f"\n✅ Final file saved to: {output_path}")


# In[34]:


import pdfplumber
import re
import pandas as pd
import os

# -------- CONFIG --------
FOLDER_PATH = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Vontobel_Files"

all_trades = []

# -------- LOOP THROUGH FILES --------
for file_name in os.listdir(FOLDER_PATH):

    if file_name.endswith(".pdf"):

        pdf_path = os.path.join(FOLDER_PATH, file_name)

        print(f"\n📄 Processing: {file_name}")

        all_text = ""
        transaction_text = ""

        # -------- READ PDF --------
        with pdfplumber.open(pdf_path) as pdf:

            for i, page in enumerate(pdf.pages):

                text = page.extract_text()

                if text:
                    all_text += "\n" + text

                # ✅ Transaction page (page 17 → index 16)
                if i == 16:
                    transaction_text = text

        # ✅ Skip file if no transaction page
        if not transaction_text:
            print("⚠️ No transaction page found — skipped")
            continue

        # -------- ACCOUNT INFO --------

        name_match = re.search(r"Nome\s+([A-Za-z,\s]+?)\s+Succursale", all_text)
        account_name = name_match.group(1).strip() if name_match else None
        if account_name:
            account_name = account_name.replace(",", "")

        portfolio_match = re.search(r"Portafoglio\s+(\d+\.\d+)", all_text)
        account_id = portfolio_match.group(1) if portfolio_match else None
        if account_id:
            account_id = account_id.split(".")[0]

        broker = "Vontobel"

        # -------- SPLIT ----------
        lines = [l.strip() for l in transaction_text.split("\n") if l.strip()]

        # -------- BUILD BLOCKS ----------
        blocks = []
        current_block = ""

        for line in lines:
            if re.match(r"^-?\s*\d[\d’.,]*", line):
                if current_block:
                    blocks.append(current_block)
                current_block = line
            else:
                current_block += " " + line

        if current_block:
            blocks.append(current_block)

        # -------- EXTRACT ----------
        for block in blocks:

            if any(x in block for x in ["Acquisto", "Vendita", "Rimborso"]):

                # ✅ TYPE
                if "Acquisto" in block:
                    trade_type = "Buy"
                elif "Vendita" in block:
                    trade_type = "Sell"
                elif "Rimborso" in block:
                    trade_type = "Sell"
                else:
                    continue

                # ✅ DATE → dd-mm-yy
                dates = re.findall(r"\d{2}\.\d{2}\.\d{4}", block)
                if dates:
                    d, m, y = dates[0].split(".")
                    trade_date = f"{d}-{m}-{y[-2:]}"
                else:
                    trade_date = None

                # ✅ QUANTITY clean
                qty_match = re.match(r"^-?\s*(\d[\d’.,]*)", block)
                quantity = qty_match.group(1) if qty_match else None
                if quantity:
                    quantity = quantity.replace("’", "").replace("'", "")

                # ✅ SECURITY clean
                security = None
                isin_match = re.search(r"[A-Z]{2}[A-Z0-9]{10}", block)

                if isin_match:
                    before_isin = block[:isin_match.start()]
                    before_isin = re.sub(r"^-?\s*\d[\d’.,]*\s+", "", before_isin)

                    sec_match = re.search(r"(.+?\b(?:EUR|CHF|USD))\b", before_isin)
                    if sec_match:
                        security = sec_match.group(1).strip()
                    else:
                        security = before_isin.strip()

                # ✅ ✅ NEW: CURRENCY EXTRACTION ✅ ✅
                currency_matches = re.findall(r"\b(EUR|CHF|USD)\b", block)
                currency = currency_matches[0] if currency_matches else None

                all_trades.append({
                    "Account Name": account_name,
                    "Account ID": account_id,
                    "Trade Date": trade_date,
                    "Type": trade_type,
                    "Security": security,
                    "Quantity": quantity,
                    "Currency": currency,   # ✅ NEW COLUMN
                    "Broker": broker
                })

# -------- FINAL OUTPUT ----------
df = pd.DataFrame(all_trades)

# ✅ Remove empty rows
if not df.empty:
    df = df[df["Security"].notna()]

print("\n✅ TOTAL TRADES EXTRACTED:", len(df))
print(df.head())

# -------- SAVE ----------
output_path = os.path.join(FOLDER_PATH, "Vontobel_trades.xlsx")

df.to_excel(output_path, index=False)

print(f"\n✅ Final file saved to: {output_path}")


# In[6]:


# add ISIN extraction
import pdfplumber
import re
import pandas as pd
import os

# -------- CONFIG --------
FOLDER_PATH = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Vontobel_Files"

all_trades = []

# -------- LOOP THROUGH FILES --------
for file_name in os.listdir(FOLDER_PATH):

    if file_name.endswith(".pdf"):

        pdf_path = os.path.join(FOLDER_PATH, file_name)

        print(f"\n📄 Processing: {file_name}")

        all_text = ""
        transaction_text = ""

        # -------- READ PDF --------
        with pdfplumber.open(pdf_path) as pdf:

            for i, page in enumerate(pdf.pages):

                text = page.extract_text()

                if text:
                    all_text += "\n" + text

                # ✅ Transaction page (page 17 → index 16)
                if i == 16:
                    transaction_text = text

        # ✅ Skip file if no transaction page
        if not transaction_text:
            print("⚠️ No transaction page found — skipped")
            continue

        # -------- ACCOUNT INFO --------

        name_match = re.search(r"Nome\s+([A-Za-z,\s]+?)\s+Succursale", all_text)
        account_name = name_match.group(1).strip() if name_match else None
        if account_name:
            account_name = account_name.replace(",", "")

        portfolio_match = re.search(r"Portafoglio\s+(\d+\.\d+)", all_text)
        account_id = portfolio_match.group(1) if portfolio_match else None
        if account_id:
            account_id = account_id.split(".")[0]

        broker = "Vontobel"

        # -------- SPLIT ----------
        lines = [l.strip() for l in transaction_text.split("\n") if l.strip()]

        # -------- BUILD BLOCKS ----------
        blocks = []
        current_block = ""

        for line in lines:
            if re.match(r"^-?\s*\d[\d’.,]*", line):
                if current_block:
                    blocks.append(current_block)
                current_block = line
            else:
                current_block += " " + line

        if current_block:
            blocks.append(current_block)

        # -------- EXTRACT ----------
        for block in blocks:

            if any(x in block for x in ["Acquisto", "Vendita", "Rimborso"]):

                # ✅ TYPE
                if "Acquisto" in block:
                    trade_type = "Buy"
                elif "Vendita" in block:
                    trade_type = "Sell"
                elif "Rimborso" in block:
                    trade_type = "Sell"
                else:
                    continue

                # ✅ DATE → dd-mm-yy
                dates = re.findall(r"\d{2}\.\d{2}\.\d{4}", block)
                if dates:
                    d, m, y = dates[0].split(".")
                    trade_date = f"{d}-{m}-{y[-2:]}"
                else:
                    trade_date = None

                # ✅ QUANTITY clean
                qty_match = re.match(r"^-?\s*(\d[\d’.,]*)", block)
                quantity = qty_match.group(1) if qty_match else None
                if quantity:
                    quantity = quantity.replace("’", "").replace("'", "")

                # ✅ ✅ ISIN EXTRACTION ✅ ✅
                isin = None
                isin_match = re.search(r"[A-Z]{2}[A-Z0-9]{10}", block)
                if isin_match:
                    isin = isin_match.group(0)

                # ✅ SECURITY clean (based on ISIN position)
                security = None
                if isin_match:
                    before_isin = block[:isin_match.start()]
                    before_isin = re.sub(r"^-?\s*\d[\d’.,]*\s+", "", before_isin)

                    sec_match = re.search(r"(.+?\b(?:EUR|CHF|USD))\b", before_isin)
                    if sec_match:
                        security = sec_match.group(1).strip()
                    else:
                        security = before_isin.strip()

                # ✅ CURRENCY
                currency_matches = re.findall(r"\b(EUR|CHF|USD)\b", block)
                currency = currency_matches[0] if currency_matches else None

                all_trades.append({
                    "Account Name": account_name,
                    "Account ID": account_id,
                    "Trade Date": trade_date,
                    "Type": trade_type,
                    "Security": security,
                    "ISIN": isin,              # ✅ NEW COLUMN
                    "Quantity": quantity,
                    "Currency": currency,
                    "Broker": broker
                })

# -------- FINAL OUTPUT ----------
df = pd.DataFrame(all_trades)

# ✅ Remove empty rows
if not df.empty:
    df = df[df["Security"].notna()]

print("\n✅ TOTAL TRADES EXTRACTED:", len(df))
print(df.head())

# -------- SAVE ----------
output_path = os.path.join(FOLDER_PATH, "Vontobel_trades.xlsx")

df.to_excel(output_path, index=False)

print(f"\n✅ Final file saved to: {output_path}")


# In[10]:


import pandas as pd
import os
import re

# ✅ Broker file paths
files = {
    "HSBC": r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\HSBC_Files\HSBC_trades.xlsx",
    "FUTU": r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Futu_Files\Futu_trades.xlsx",
    "IBKR": r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\IBKR_Files\IBKR_trades.xlsx",
    "Vontobel": r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Vontobel_Files\Vontobel_trades.xlsx"
}

# ✅ Output file
output_file = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Combined_Broker_Trades.xlsx"

# ✅ Function: detect Chinese characters
def contains_chinese(text):
    if isinstance(text, str):
        return bool(re.search(r'[\u4e00-\u9fff]', text))
    return False


all_dfs = []

# ---- LOOP EACH BROKER ----
for broker, file_path in files.items():

    print(f"Processing: {broker}")

    try:
        df = pd.read_excel(file_path, dtype={"Account ID": str})

        # ✅ FUTU handling
        if broker == "FUTU" and "Securities (Eng)" in df.columns:

            original_security = df["Security"]
            df["Security"] = df["Securities (Eng)"]

            df["Security(cn)"] = original_security.apply(
                lambda x: x if contains_chinese(x) else None
            )

        else:
            df["Security(cn)"] = None

        # ✅ Ensure required columns exist
        if "Symbol" not in df.columns:
            df["Symbol"] = None

        if "ISIN" not in df.columns:
            df["ISIN"] = None

        if "Currency" not in df.columns:
            df["Currency"] = None

        # ✅ Add Broker
        df["Broker"] = broker

        # ✅ Keep only required columns (now safe)
        df = df[
            ["Account Name", "Account ID", "Trade Date",
             "Type", "Security", "Security(cn)", "Quantity",
             "Currency", "Symbol", "ISIN", "Broker"]
        ]

        all_dfs.append(df)

    except Exception as e:
        print(f"Error reading {broker}: {e}")

# ---- COMBINE ----
combined_df = pd.concat(all_dfs, ignore_index=True)

# ✅ CLEAN Trade Date
combined_df["Trade Date"] = combined_df["Trade Date"].astype(str).str.strip()

# ✅ Fix Vontobel 2-digit year
def fix_vontobel_date(x):
    if isinstance(x, str) and re.match(r"\d{2}-\d{2}-\d{2}$", x):
        d, m, y = x.split("-")
        return f"{d}-{m}-20{y}"
    return x

combined_df["Trade Date"] = combined_df["Trade Date"].apply(fix_vontobel_date)

# ✅ Convert to datetime
combined_df["Trade Date"] = pd.to_datetime(
    combined_df["Trade Date"],
    dayfirst=True,
    errors="coerce"
)

# ✅ Sort
combined_df = combined_df.sort_values(by=["Trade Date", "Account Name"])

# ✅ Convert back to string
combined_df["Trade Date"] = combined_df["Trade Date"].dt.strftime("%d-%m-%Y")

# ✅ Replace NaN with None (optional cleaner output)
combined_df = combined_df.where(pd.notnull(combined_df), None)

# ✅ Save
combined_df.to_excel(output_file, index=False)

print("\n✅ DONE — Combined file saved to:")
print(output_file)
print(combined_df)


# # Combine all records from different brokers

# In[18]:


import pandas as pd
import os
import re

# ✅ Broker file paths
files = {
    "HSBC": r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\HSBC_Files\HSBC_trades.xlsx",
    "FUTU": r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Futu_Files\Futu_trades.xlsx",
    "IBKR": r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\IBKR_Files\IBKR_trades.xlsx",
    "Vontobel": r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Vontobel_Files\Vontobel_trades.xlsx"
}

# ✅ Output file
output_file = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Combined_Broker_Trades.xlsx"

# ✅ Detect Chinese
def contains_chinese(text):
    if isinstance(text, str):
        return bool(re.search(r'[\u4e00-\u9fff]', text))
    return False


all_dfs = []

# ---- LOOP ----
for broker, file_path in files.items():

    print(f"Processing: {broker}")

    try:
        df = pd.read_excel(file_path, dtype={"Account ID": str})

        # =========================
        # ✅ FUTU handling (FIXED ✅)
        # =========================
        if broker == "FUTU":

            # ✅ Ensure columns exist
            if "ISIN" not in df.columns:
                df["ISIN"] = None

            if "Symbol" not in df.columns:
                df["Symbol"] = None

            # ✅ Preserve original Chinese name
            if "Securities (Eng)" in df.columns:
                original_security = df["Security"]

                df["Security"] = df["Securities (Eng)"]

                df["Security(cn)"] = original_security.apply(
                    lambda x: x if contains_chinese(x) else None
                )
            else:
                df["Security(cn)"] = None

            # ✅ IMPORTANT: keep ISIN vs Symbol logic
            df["ISIN"] = df["ISIN"].where(df["ISIN"].notna(), None)
            df["Symbol"] = df["Symbol"].where(df["Symbol"].notna(), None)

        else:
            df["Security(cn)"] = None

        # =========================
        # ✅ Ensure required columns
        # =========================
        for col in ["Symbol", "ISIN", "Currency"]:
            if col not in df.columns:
                df[col] = None

        # ✅ Add broker
        df["Broker"] = broker

        # ✅ Keep consistent columns
        df = df[
            ["Account Name", "Account ID", "Trade Date",
             "Type", "Security", "Security(cn)",
             "Quantity", "Currency",
             "Symbol", "ISIN", "Broker"]
        ]

        all_dfs.append(df)

    except Exception as e:
        print(f"Error reading {broker}: {e}")

# =========================
# ✅ COMBINE
# =========================
combined_df = pd.concat(all_dfs, ignore_index=True)

# =========================
# ✅ CLEAN DATE
# =========================
combined_df["Trade Date"] = combined_df["Trade Date"].astype(str).str.strip()

# ✅ Fix Vontobel dates
def fix_vontobel_date(x):
    if isinstance(x, str) and re.match(r"\d{2}-\d{2}-\d{2}$", x):
        d, m, y = x.split("-")
        return f"{d}-{m}-20{y}"
    return x

combined_df["Trade Date"] = combined_df["Trade Date"].apply(fix_vontobel_date)

# ✅ Convert to datetime
combined_df["Trade Date"] = pd.to_datetime(
    combined_df["Trade Date"],
    dayfirst=True,
    errors="coerce"
)

# ✅ Sort
combined_df = combined_df.sort_values(by=["Trade Date", "Account Name"])

# ✅ Back to string
combined_df["Trade Date"] = combined_df["Trade Date"].dt.strftime("%d-%m-%Y")

# ✅ Clean NaN
combined_df = combined_df.where(pd.notnull(combined_df), None)

# =========================
# ✅ SAVE
# =========================
combined_df.to_excel(output_file, index=False)

print("\n✅ DONE — Combined file saved to:")
print(output_file)
print(combined_df)


# In[19]:


# add staff name column
import pandas as pd

# ✅ Combined file (your output)
combined_file = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Combined_Broker_Trades.xlsx"

# ✅ Mapping file
mapping_file = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Staff Name and Account Name.xlsx"

# ✅ Output file
output_file = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Combined_Broker_Trades.xlsx"


# ---- LOAD FILES ----
df = pd.read_excel(combined_file)
mapping = pd.read_excel(mapping_file)


# ✅ Clean column names (very important)
mapping.columns = mapping.columns.str.strip()
df.columns = df.columns.str.strip()


# ---- CREATE MAPPING DICTIONARY ----
# Key: Statements name → Value: CIGP name
name_map = dict(zip(
    mapping["Staff Names as per Statements"],
    mapping["Staff Names as per CIGP"]
))


# ---- MAP STAFF NAME ----
df["Staff Name"] = df["Account Name"].map(name_map)


# ✅ If no match → keep original (optional but recommended)
df["Staff Name"] = df["Staff Name"].fillna(df["Account Name"])


# ---- REORDER COLUMNS (put Staff Name next to Account Name) ----
cols = df.columns.tolist()

# Remove Staff Name and reinsert after Account Name
cols.remove("Staff Name")
account_index = cols.index("Account Name")
cols.insert(account_index + 1, "Staff Name")

df = df[cols]


# ---- SAVE OUTPUT ----
df.to_excel(output_file, index=False)


print("\n✅ DONE — Staff Name added successfully")
print("Saved to:", output_file)
print(df)


# In[20]:


# add staff name column
import pandas as pd
from rapidfuzz import fuzz


# =========================
# ✅ Files
# =========================
combined_file = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Combined_Broker_Trades.xlsx"
mapping_file = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Staff Name and Account Name.xlsx"
output_file = combined_file


# =========================
# ✅ Load
# =========================
df = pd.read_excel(combined_file)
mapping = pd.read_excel(mapping_file)

mapping.columns = mapping.columns.str.strip()
df.columns = df.columns.str.strip()


# =========================
# ✅ Normalize function
# =========================
def normalize(text):
    return (
        str(text)
        .lower()
        .replace(" ", "")
        .replace(".", "")
        .replace(",", "")
    )


# =========================
# ✅ Build mapping list
# =========================
mapping_list = list(zip(
    mapping["Staff Names as per Statements"],
    mapping["Staff Names as per CIGP"]
))


# =========================
# ✅ Improved matching function
# =========================
def match_staff(account_name):

    account_norm = normalize(account_name)

    best_score = 0
    best_staff = None

    for stmt_name, cigp_name in mapping_list:

        stmt_norm = normalize(stmt_name)

        score = fuzz.ratio(account_norm, stmt_norm)

        if score > best_score:
            best_score = score
            best_staff = cigp_name

    # ✅ threshold to avoid wrong match
    if best_score >= 80:
        return best_staff
    else:
        return account_name   # fallback


# =========================
# ✅ Apply matching
# =========================
df["Staff Name"] = df["Account Name"].apply(match_staff)


# =========================
# ✅ Reorder columns
# =========================
cols = df.columns.tolist()

cols.remove("Staff Name")
account_index = cols.index("Account Name")
cols.insert(account_index + 1, "Staff Name")

df = df[cols]


# =========================
# ✅ Save
# =========================
df.to_excel(output_file, index=False)


print("\n✅ DONE — Robust staff matching applied")
print("Saved to:", output_file)


# In[21]:


# clean symbol and ISIN
import pandas as pd
import re

# =========================
# 1. LOAD FILE
# =========================
file_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Combined_Broker_Trades.xlsx"   # 👈 update path
df = pd.read_excel(file_path)

# =========================
# 2. CLEAN FUNCTION
# =========================
def clean_symbol(x):
    if pd.isna(x):
        return x

    x = str(x).strip()

    # ✅ Case 1: HK.HK0000502390 → HK0000502390
    if x.startswith("HK.HK"):
        return x.replace("HK.", "", 1)

    # ✅ Case 2: HK.03690 → 3690 HK
    if x.startswith("HK."):
        num = x.split("HK.")[1]
        num = re.sub(r"^0+", "", num)
        return f"{num} HK"

    return x

# =========================
# 3. APPLY (OVERWRITE)
# =========================
df["Symbol"] = df["Symbol"].apply(clean_symbol)
if "Symbol_clean" in df.columns:
    df = df.drop(columns=["Symbol_clean"])

# =========================
# 4. SAVE
# =========================
output_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Combined_Broker_Trades.xlsx"
df.to_excel(output_path, index=False)

print("✅ Symbol column updated successfully")


# # Merge with security code file to match ISIN and symbol

# In[22]:


get_ipython().system('pip install rapidfuzz')


# In[23]:


import pandas as pd
import re
from rapidfuzz import process, fuzz

# =========================
# 1. LOAD FILES
# =========================
file_main = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Combined_Broker_Trades.xlsx"
file_ref = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Securities_20260624.xlsx"

df = pd.read_excel(file_main)
ref = pd.read_excel(file_ref)

# =========================
# 2. STANDARDIZE REFERENCE
# =========================
ref = ref.rename(columns={
    "Security code": "symbol_ref",
    "Security ISIN code": "isin_ref",
    "Security name": "name_ref"
})

# clean text
def clean_text(x):
    if pd.isna(x):
        return ""
    x = str(x).lower().strip()
    x = re.sub(r"\s+", " ", x)
    return x

ref["name_ref_clean"] = ref["name_ref"].apply(clean_text)

# =========================
# 3. PREP MAIN DATA
# =========================
df["Security_clean"] = df["Security"].apply(clean_text)

# =========================
# 4. BUILD LOOKUPS
# =========================
symbol_to_isin = dict(zip(ref["symbol_ref"], ref["isin_ref"]))
isin_to_symbol = dict(zip(ref["isin_ref"], ref["symbol_ref"]))

names_list = ref["name_ref_clean"].tolist()

# =========================
# 5. FUZZY MATCH FUNCTION
# =========================
def fuzzy_match_name(name):
    if name == "":
        return None, None

    match = process.extractOne(
        name,
        names_list,
        scorer=fuzz.token_sort_ratio
    )

    if match and match[1] >= 80:   # ✅ threshold adjustable
        matched_name = match[0]
        matched_row = ref[ref["name_ref_clean"] == matched_name].iloc[0]
        return matched_row["symbol_ref"], matched_row["isin_ref"]

    return None, None

# =========================
# 6. ENRICH DATA
# =========================
new_symbols = []
new_isins = []

for _, row in df.iterrows():

    symbol = row.get("Symbol")
    isin = row.get("ISIN")
    name = row.get("Security_clean")

    new_symbol = symbol
    new_isin = isin

    # =========================
    # ✅ CASE 1: Missing ISIN
    # =========================
    if pd.isna(isin) or str(isin).strip() == "":

        if pd.notna(symbol) and symbol in symbol_to_isin:
            new_isin = symbol_to_isin[symbol]

        else:
            sym, isin_val = fuzzy_match_name(name)
            if isin_val:
                new_isin = isin_val
                if not new_symbol:
                    new_symbol = sym

    # =========================
    # ✅ CASE 2: Missing SYMBOL
    # =========================
    if pd.isna(symbol) or str(symbol).strip() == "":

        if pd.notna(isin) and isin in isin_to_symbol:
            new_symbol = isin_to_symbol[isin]

        else:
            sym, isin_val = fuzzy_match_name(name)
            if sym:
                new_symbol = sym
                if not new_isin:
                    new_isin = isin_val

    new_symbols.append(new_symbol)
    new_isins.append(new_isin)

# =========================
# 7. UPDATE ORIGINAL COLUMNS
# =========================
df["Symbol"] = new_symbols
df["ISIN"] = new_isins

# =========================
# 8. SAVE
# =========================
output_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Combined_Broker_Trades_ISIN.xlsx"
df.to_excel(output_path, index=False)

print("✅ Matching completed: Symbol & ISIN enriched")


# In[24]:


import pandas as pd
import re
from rapidfuzz import process, fuzz

# =========================
# 1. LOAD FILES
# =========================
file_main = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Combined_Broker_Trades.xlsx"
file_ref = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Securities_20260624.xlsx"

df = pd.read_excel(file_main)
ref = pd.read_excel(file_ref)


# =========================
# 2. STANDARDIZE REFERENCE
# =========================
ref = ref.rename(columns={
    "Security code": "symbol_ref",
    "Security ISIN code": "isin_ref",
    "Security name": "name_ref"
})

# Build lookup maps
symbol_to_isin = dict(zip(ref["symbol_ref"], ref["isin_ref"]))
isin_to_symbol = dict(zip(ref["isin_ref"], ref["symbol_ref"]))

# Clean helper
def clean_text(x):
    if pd.isna(x):
        return ""
    return re.sub(r"\s+", " ", str(x).lower().strip())

# Pre-create name list
ref["name_clean"] = ref["name_ref"].apply(clean_text)
name_list = ref["name_clean"].tolist()

# =========================
# 3. PASS 1 — ISIN → SYMBOL
# =========================
for i, row in df.iterrows():
    isin = row["ISIN"]
    symbol = row["Symbol"]

    if (pd.notna(isin) and isin in isin_to_symbol) and (pd.isna(symbol) or symbol == ""):
        df.at[i, "Symbol"] = isin_to_symbol[isin]

# =========================
# 4. PASS 2 — SYMBOL → ISIN
# =========================
for i, row in df.iterrows():
    symbol = row["Symbol"]
    isin = row["ISIN"]

    if (pd.notna(symbol) and symbol in symbol_to_isin) and (pd.isna(isin) or isin == ""):
        df.at[i, "ISIN"] = symbol_to_isin[symbol]

# =========================
# 5. PASS 3 — FUZZY MATCH (LAST STEP ONLY)
# =========================
def fuzzy_match(name):
    if name == "":
        return None, None

    match = process.extractOne(name, name_list, scorer=fuzz.token_sort_ratio)

    if match and match[1] >= 85:   # stricter threshold
        matched = ref[ref["name_clean"] == match[0]].iloc[0]
        return matched["symbol_ref"], matched["isin_ref"]

    return None, None

for i, row in df.iterrows():
    symbol = row["Symbol"]
    isin = row["ISIN"]

    # only run if still missing something
    if pd.isna(symbol) or pd.isna(isin):

        name_clean = clean_text(row["Security"])

        sym, isin_val = fuzzy_match(name_clean)

        if pd.isna(symbol) and sym:
            df.at[i, "Symbol"] = sym

        if pd.isna(isin) and isin_val:
            df.at[i, "ISIN"] = isin_val

# =========================
# 6. SAVE
# =========================
df.to_excel(r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Combined_Broker_Trades_ISIN.xlsx", index=False)

print("✅ Matching completed — clean & accurate")


# In[29]:


import pandas as pd
import re
from rapidfuzz import process, fuzz

# =========================
# 1. LOAD
# =========================
file_main = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Combined_Broker_Trades.xlsx"
file_ref = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Securities_20260624.xlsx"

df = pd.read_excel(file_main)
ref = pd.read_excel(file_ref)

# =========================
# 2. STANDARDIZE REFERENCE
# =========================
ref = ref.rename(columns={
    "Security code": "symbol_ref",
    "Security ISIN code": "isin_ref",
    "Security trade code": "trade_code_ref",
    "Security name": "name_ref"
})

# =========================
# 3. NORMALIZATION FUNCTIONS
# =========================
def clean_name(x):
    if pd.isna(x):
        return ""

    x = str(x).lower()

    # remove brackets
    x = re.sub(r"\(.*?\)", "", x)

    # remove punctuation
    x = re.sub(r"[,.-]", " ", x)

    x = re.sub(r"\s+", " ", x).strip()

    return x


def normalize_symbol(x):
    if pd.isna(x):
        return ""

    x = str(x).upper().strip()
    x = x.replace("EQUITY", "")
    x = x.replace("-", "")
    x = x.replace(" ", "")

    return x

# =========================
# 4. PREP REFERENCE DATA
# =========================
ref["name_clean"] = ref["name_ref"].apply(clean_name)
ref["symbol_norm"] = ref["symbol_ref"].apply(normalize_symbol)
ref["trade_norm"] = ref["trade_code_ref"].apply(normalize_symbol)

name_list = ref["name_clean"].tolist()

# =========================
# 5. LOOKUPS (ONLY FOR ISIN STEP)
# =========================
symbol_to_isin = dict(zip(ref["symbol_norm"], ref["isin_ref"]))
trade_to_isin = dict(zip(ref["trade_norm"], ref["isin_ref"]))

# =========================
# ✅ STEP 1 — FILL SYMBOL (ONLY via NAME)
# =========================
print("\n===== STEP 1: SYMBOL MATCH =====")

def match_symbol_by_name(security_name):

    name_clean = clean_name(security_name)

    match = process.extractOne(
        name_clean,
        name_list,
        scorer=fuzz.partial_ratio
    )

    if match is None:
        return None

    score = match[1]

    # ✅ strict threshold to avoid wrong matches
    if score < 85:
        return None

    matched_row = ref[ref["name_clean"] == match[0]].iloc[0]

    return matched_row["symbol_ref"]

# apply ONLY for missing symbol
for i, row in df.iterrows():

    symbol = row["Symbol"]

    if pd.isna(symbol) or str(symbol).strip() == "":
        matched_symbol = match_symbol_by_name(row["Security"])

        if matched_symbol:
            df.at[i, "Symbol"] = matched_symbol
            print(f"[SYMBOL MATCH] {row['Security']} → {matched_symbol}")

print("===== STEP 1 DONE =====\n")

# =========================
# ✅ STEP 2 — FILL ISIN (ONLY via SYMBOL)
# =========================
print("\n===== STEP 2: ISIN MATCH =====")

for i, row in df.iterrows():

    symbol_raw = row["Symbol"]
    isin = row["ISIN"]

    if pd.notna(symbol_raw) and (pd.isna(isin) or isin == ""):

        symbol_norm = normalize_symbol(symbol_raw)

        # ✅ PRIORITY 1: Security code
        if symbol_norm in symbol_to_isin:
            df.at[i, "ISIN"] = symbol_to_isin[symbol_norm]
            print(f"[ISIN CODE] {symbol_raw}")

        # ✅ PRIORITY 2: Trade code
        elif symbol_norm in trade_to_isin:
            df.at[i, "ISIN"] = trade_to_isin[symbol_norm]
            print(f"[ISIN TRADE] {symbol_raw}")

        # ✅ PRIORITY 3: US stock format
        else:
            us_candidate = symbol_norm + "US"

            if us_candidate in trade_to_isin:
                df.at[i, "ISIN"] = trade_to_isin[us_candidate]
                print(f"[ISIN US] {symbol_raw} → {us_candidate}")

        # ❗ NO FUZZY HERE (IMPORTANT)

print("===== STEP 2 DONE =====\n")

# =========================
# 6. SAVE
# =========================
output_file = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Combined_Broker_Trades_ISIN.xlsx"
df.to_excel(output_file, index=False)

output_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\PAD sheets.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
   df.to_excel(writer, sheet_name="Extraction from Statement", index=False)

print("✅ FINAL — SAFE MATCHING COMPLETED")


# # Match with PAD protal records

# In[26]:


import pandas as pd
from rapidfuzz import fuzz
from pandas.tseries.offsets import BDay
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import yfinance as yf
import re


# =========================
# ✅ Load data
# =========================
trade_df = pd.read_excel(r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Combined_Broker_Trades.xlsx")
pad_df = pd.read_excel(r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\Export from PAD Portal.xlsx")

trade_df["Trade Date"] = pd.to_datetime(trade_df["Trade Date"], dayfirst=True)
pad_df["Request Date"] = pd.to_datetime(pad_df["Request Date"], dayfirst=True)


# =========================
# ✅ Normalize + enhance
# =========================
def normalize(text):
    text = str(text).lower()
    remove_words = ["limited","ltd","inc","corp","holdings","group","class","co","shs"]
    for w in remove_words:
        text = text.replace(w,"")
    return " ".join(text.replace("-", " ").split())


def enhance_name(text):
    text = normalize(text)
    replacements = {
        "jingdong": "jd",
        "jian kang": "health",
        "qi che": "auto",
        "lanyueliang": "blue moon",
        "lan yue liang": "blue moon",
        "mei tuan": "meituan"
    }
    for k,v in replacements.items():
        text = text.replace(k,v)
    return text


# =========================
# ✅ ETF detection (IMPROVED)
# =========================
def is_etf(security):
    sec = str(security).lower()
    return (
        "etf" in sec or
        "index" in sec or
        "nasdaq" in sec or
        "track" in sec or
        "pre" in sec
    )


# ✅ Better ticker extraction
def extract_ticker(text):

    words = str(text).split()

    for w in words:
        clean = re.sub(r'[^A-Z0-9]', '', w)

        if (
            clean.isalnum() and
            clean.isupper() and
            2 <= len(clean) <= 6 and
            clean not in ["SHS", "USD", "HKD", "ORD"]
        ):
            return clean

    return None

# ✅ ETF AUM cache
ETF_CACHE = {}

def get_etf_aum(ticker):

    if ticker in ETF_CACHE:
        return ETF_CACHE[ticker]

    try:
        info = yf.Ticker(ticker).info
        aum = info.get("totalAssets", None)
        ETF_CACHE[ticker] = aum
        return aum
    except:
        ETF_CACHE[ticker] = None
        return None


# ✅ ETF status
def check_etf_status(security):

    sec = str(security)

    # ✅ Step 1: Known ETF keywords (STRICT)
    ETF_KEYWORDS = ["ETF"]

    if any(k in sec.upper() for k in ETF_KEYWORDS):
        keyword_flag = True
    else:
        keyword_flag = False

    # ✅ Step 2: Extract ticker safely
    ticker = extract_ticker(sec)

    if not ticker:
        return "NOT_ETF", None, None

    try:
        info = yf.Ticker(ticker).info
        quote_type = info.get("quoteType", None)

        # ✅ Step 3: Only trust Yahoo ETF label
        if quote_type != "ETF":
            return "NOT_ETF", ticker, None

        # ✅ Step 4: AUM check
        aum = info.get("totalAssets", None)

        if aum is None:
            return "ETF_UNCERTAIN", ticker, None

        if aum > 200_000_000:
            return "ETF_EXEMPT", ticker, aum
        else:
            return "ETF_CONTROL", ticker, aum

    except:
        # ✅ ONLY fallback if strong keyword signal exists
        if keyword_flag:
            return "ETF_UNCERTAIN", ticker, None
        else:
            return "NOT_ETF", ticker, None


# =========================
# ✅ Late trade rule
# =========================
def is_late_trade(req_date, trade_date):
    return trade_date > req_date + BDay(3)


# =========================
# ✅ Security selection
# =========================
def get_trade_security(row):
    if "Securities (Eng)" in trade_df.columns:
        return row["Securities (Eng)"]
    return row["Security"]


# =========================
# ✅ MAIN PROCESS
# =========================
results = []

for _, trade in trade_df.iterrows():

    staff = trade["Staff Name"]
    trade_date = trade["Trade Date"]
    trade_sec = get_trade_security(trade)
    trade_type = trade["Type"]
    trade_qty = trade["Quantity"]

    issues = []

    # ✅ ETF check
    etf_status, ticker, aum = check_etf_status(trade_sec)

    if etf_status == "ETF_EXEMPT":
        results.append({
            **trade,
            "Matched Security": "ETF Exempt",
            "Match Score": None,
            "Request Date": None,
            "Approved Qty": None,
            "ETF Status": etf_status,
            "AUM": aum,
            "Ticker": ticker,
            "Unmatched Approval Security": "",
            "Unmatched Approval Security Request Qty": "",
            "Issue": "OK (ETF >200M Exempt)"
        })
        continue

    if etf_status == "ETF_UNCERTAIN":
        issues.append("ETF AUM Uncertain")

    # ✅ PAD matching
    pad_filtered = pad_df[pad_df["Requestor"] == staff].copy()
    pad_filtered = pad_filtered[pad_filtered["Request Date"] <= trade_date]

    if pad_filtered.empty:
        results.append({**trade, "Issue": "No prior PAD approval"})
        continue

    pad_filtered["Date Diff"] = (trade_date - pad_filtered["Request Date"]).dt.days
    pad_filtered = pad_filtered.sort_values("Date Diff").head(5)

    best_score = 0
    best_row = None

    for _, p in pad_filtered.iterrows():
        score = fuzz.token_sort_ratio(
            enhance_name(trade_sec),
            enhance_name(p["Security"])
        )

        if score > best_score:
            best_score = score
            best_row = p

    if best_score < 70:
        issues.append("Low Match Score")

    if best_score < 50 or best_row is None:
        issues.append("Unapproved Security")

        results.append({
            **trade,
            "Matched Security": None,
            "Match Score": best_score,
            "ETF Status": etf_status,
            "AUM": aum,
            "Ticker": ticker,
            "Unmatched Approval Security": "",
            "Unmatched Approval Security Request Qty": "",
            "Issue": "; ".join(issues)
        })
        continue

    req_date = best_row["Request Date"]
    req_type = best_row["Type"]
    req_qty = best_row["Quantity"]

    if is_late_trade(req_date, trade_date):
        issues.append("Late Trade")

    if str(trade_type).lower() != str(req_type).lower():
        issues.append("Type Mismatch")

    if trade_qty > req_qty:
        issues.append("Quantity Exceeded")

    results.append({
        **trade,
        "Matched Security": best_row["Security"],
        "Match Score": best_score,
        "Request Date": req_date,
        "Approved Qty": req_qty,
        "ETF Status": etf_status,
        "AUM": aum,
        "Ticker": ticker,
        "Unmatched Approval Security": "",
        "Unmatched Approval Security Request Qty": "",
        "Issue": "OK" if not issues else "; ".join(issues)
    })


df = pd.DataFrame(results)


# =========================
# ✅ Unmatched approvals
# =========================
for key, group in df.groupby(["Staff Name", "Trade Date"]):

    staff, date = key

    pad_group = pad_df[
        (pad_df["Requestor"] == staff) &
        (pad_df["Request Date"] == date)
    ]

    pad_dict = dict(zip(pad_group["Security"], pad_group["Quantity"]))
    matched = set(group["Matched Security"].dropna())

    unmatched = set(pad_dict.keys()) - matched

    df.loc[group.index, "Unmatched Approval Security"] = ", ".join(unmatched)
    df.loc[group.index, "Unmatched Approval Security Request Qty"] = ", ".join(
        f"{k}: {pad_dict[k]}" for k in unmatched
    )

# =========================
# ✅ Format dates (REMOVE TIME)
# =========================
if "Trade Date" in df.columns:
    df["Trade Date"] = df["Trade Date"].dt.strftime("%d-%m-%Y")

if "Request Date" in df.columns:
    df["Request Date"] = pd.to_datetime(df["Request Date"], errors="coerce").dt.strftime("%d-%m-%Y")
# =========================
# ✅ SAVE
# =========================
df = df.drop(columns=["AUM", "Ticker", "ETF Status"], errors="ignore")
output_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\PAD_Control_Result_original.xlsx"
df.to_excel(output_path, index=False)


# =========================
# ✅ Highlight
# =========================
wb = load_workbook(output_path)
ws = wb.active

red = PatternFill(start_color="FF9999", fill_type="solid")
yellow = PatternFill(start_color="FFFF99", fill_type="solid")

issue_col = [c.value for c in ws[1]].index("Issue") + 1

for r in range(2, ws.max_row+1):
    val = ws.cell(r, issue_col).value

    if val:
        if "Unapproved Security" in val:
            for c in range(1, ws.max_column+1):
                ws.cell(r, c).fill = red

        elif "ETF AUM Uncertain" in val:
            for c in range(1, ws.max_column+1):
                ws.cell(r, c).fill = yellow


wb.save(output_path)

print("✅ FINAL VERSION COMPLETE")


# In[35]:


import pandas as pd
from rapidfuzz import fuzz
from pandas.tseries.offsets import BDay
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import yfinance as yf
import re

# =========================
# ✅ LOAD DATA
# =========================
trade_df = pd.read_excel(
    r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\PAD sheets.xlsx",
    sheet_name="Extraction from Statement"
)

pad_df = pd.read_excel(
    r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\PAD sheets.xlsx",
    sheet_name="Extraction from PAD Portal"
)

# ✅ Trade date
trade_df["Trade Date"] = pd.to_datetime(
    trade_df["Trade Date"],
    dayfirst=True,
    errors="coerce"
)

# ✅ Final Approval Date (UPDATED ✅)
pad_df["Final Approval Date"] = pd.to_datetime(
    pad_df["Final Approval Date"],
    format="%d %b %Y",
    errors="coerce"
)

# =========================
# ✅ NORMALIZATION
# =========================
def normalize(text):
    text = str(text).lower()
    remove_words = ["limited","ltd","inc","corp","holdings","group","class","co","shs"]
    for w in remove_words:
        text = text.replace(w,"")
    return " ".join(text.replace("-", " ").split())


def enhance_name(text):
    text = normalize(text)
    replacements = {
        "jingdong": "jd",
        "jian kang": "health",
        "qi che": "auto",
        "lanyueliang": "blue moon",
        "mei tuan": "meituan"
    }
    for k,v in replacements.items():
        text = text.replace(k,v)
    return text


# =========================
# ✅ ETF CHECK
# =========================
def extract_ticker(text):
    for w in str(text).split():
        clean = re.sub(r'[^A-Z0-9]', '', w)
        if clean.isupper() and 2 <= len(clean) <= 6:
            return clean
    return None


def check_etf_status(security):
    ticker = extract_ticker(security)
    if not ticker:
        return "NOT_ETF", None, None

    try:
        info = yf.Ticker(ticker).info
        if info.get("quoteType") != "ETF":
            return "NOT_ETF", ticker, None

        aum = info.get("totalAssets")

        if aum is None:
            return "ETF_UNCERTAIN", ticker, None

        if aum > 200_000_000:
            return "ETF_EXEMPT", ticker, aum

        return "ETF_CONTROL", ticker, aum

    except:
        return "NOT_ETF", ticker, None


# =========================
# ✅ DATE RULE
# =========================
def is_late_trade(req_date, trade_date):
    return trade_date > req_date + BDay(3)


# =========================
# ✅ MAIN PROCESS
# =========================
results = []

for _, trade in trade_df.iterrows():

    staff = trade["Staff Name"]
    trade_date = trade["Trade Date"]
    trade_sec = trade["Security"]
    trade_type = trade["Type"]
    trade_qty = trade["Quantity"]
    trade_isin = trade.get("ISIN")

    issues = []

    # =========================
    # ✅ FILTER PAD DATA (UPDATED DATE ✅)
    # =========================
    pad_filtered = pad_df[
        (pad_df["Requestor"] == staff) &
        (pad_df["Final Approval Date"] <= trade_date)  # ✅ updated
    ].copy()

    if pad_filtered.empty:
        results.append({**trade, "Issue": "No prior PAD approval"})
        continue

    pad_filtered["Date Diff"] = (
        trade_date - pad_filtered["Final Approval Date"]
    ).dt.days

    pad_filtered = pad_filtered.sort_values("Date Diff").head(5)

    # =========================
    # ✅ MATCHING
    # =========================
    best_row = None
    best_score = 0

    # ✅ STEP 1 — ISIN MATCH
    if pd.notna(trade_isin):

        isin_match = pad_filtered[
            pad_filtered["ISIN Code"] == trade_isin
        ]

        if not isin_match.empty:
            best_row = isin_match.iloc[0]
            best_score = 100

    # ✅ STEP 2 — FUZZY FALLBACK
    if best_row is None:

        for _, p in pad_filtered.iterrows():

            pad_isin = p.get("ISIN Code")

            if pd.notna(trade_isin) and pd.notna(pad_isin):
                continue

            score = fuzz.token_sort_ratio(
                enhance_name(trade_sec),
                enhance_name(p["Security"])
            )

            if score > best_score:
                best_score = score
                best_row = p

    # ✅ REJECT WEAK MATCH
    if best_score < 70 or best_row is None:

        results.append({
            **trade,
            "Matched Security": None,
            "Match Score": best_score,
            "Issue": "Unapproved Security"
        })
        continue

    # =========================
    # ✅ VALID MATCH
    # =========================
    req_date = best_row["Final Approval Date"]
    req_type = best_row["Type"]
    req_qty = best_row["Quantity"]

    if is_late_trade(req_date, trade_date):
        issues.append("Late Trade")

    if str(trade_type).lower() != str(req_type).lower():
        issues.append("Type Mismatch")

    if trade_qty > req_qty:
        issues.append("Quantity Exceeded")

    results.append({
        **trade,
        "Matched Security": best_row["Security"],
        "Match Score": best_score,
        "Request Date": req_date,
        "Approved Qty": req_qty,
        "Issue": "OK" if not issues else "; ".join(issues)
    })


df = pd.DataFrame(results)


# =========================
# ✅ ✅ ADD BACK UNMATCHED APPROVAL ✅
# =========================
for (staff, date), group in df.groupby(["Staff Name", "Trade Date"]):

    pad_group = pad_df[
        (pad_df["Requestor"] == staff) &
        (pad_df["Final Approval Date"] == pd.to_datetime(date))
    ]

    pad_dict = dict(zip(pad_group["Security"], pad_group["Quantity"]))
    matched = set(group["Matched Security"].dropna())

    unmatched = set(pad_dict.keys()) - matched

    df.loc[group.index, "Unmatched Approval Security"] = ", ".join(unmatched)

    df.loc[group.index, "Unmatched Approval Security Request Qty"] = ", ".join(
        f"{k}: {pad_dict[k]}" for k in unmatched
    )


# =========================
# ✅ FORMAT DATE OUTPUT
# =========================
df["Trade Date"] = pd.to_datetime(df["Trade Date"]).dt.strftime("%d-%m-%Y")

if "Request Date" in df.columns:
    df["Request Date"] = pd.to_datetime(df["Request Date"], errors="coerce").dt.strftime("%d-%m-%Y")

# =========================
# ✅ SAVE
# =========================
output_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\PAD sheets.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df.to_excel(writer, sheet_name="PAD Results", index=False)


# =========================
# ✅ HIGHLIGHT
# =========================
wb = load_workbook(output_path)
ws = wb["PAD Results"]

red = PatternFill(start_color="FF9999", fill_type="solid")

issue_col = [c.value for c in ws[1]].index("Issue") + 1

for r in range(2, ws.max_row + 1):
    val = ws.cell(r, issue_col).value

    if val and "Unapproved" in val:
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).fill = red

wb.save(output_path)

print("✅ FINAL — FULL FIXED VERSION COMPLETE")


# import pandas as pd
# from rapidfuzz import fuzz
# from pandas.tseries.offsets import BDay
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
# import yfinance as yf
# import re
# 
# # =========================
# # ✅ LOAD DATA
# # =========================
# trade_df = pd.read_excel(
#     r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\PAD sheets.xlsx",
#     sheet_name="Extraction from Statement"
# )
# 
# pad_df = pd.read_excel(
#     r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\PAD sheets.xlsx",
#     sheet_name="Extraction from PAD Portal"
# )
# 
# 
# # ✅ Dates
# trade_df["Trade Date"] = pd.to_datetime(trade_df["Trade Date"], dayfirst=True, errors="coerce")
# pad_df["Final Approval Date"] = pd.to_datetime(pad_df["Final Approval Date"], format="%d %b %Y", errors="coerce")
# 
# # =========================
# # ✅ NORMALIZATION
# # =========================
# def normalize(text):
#     text = str(text).lower()
#     remove_words = ["limited","ltd","inc","corp","holdings","group","class","co","shs"]
#     for w in remove_words:
#         text = text.replace(w,"")
#     return " ".join(text.replace("-", " ").split())
# 
# def enhance_name(text):
#     text = normalize(text)
#     replacements = {
#         "jingdong": "jd",
#         "jian kang": "health",
#         "qi che": "auto",
#         "lanyueliang": "blue moon",
#         "mei tuan": "meituan"
#     }
#     for k,v in replacements.items():
#         text = text.replace(k,v)
#     return text
# 
# # =========================
# # ✅ ETF CHECK (kept unchanged)
# # =========================
# def extract_ticker(text):
#     for w in str(text).split():
#         clean = re.sub(r'[^A-Z0-9]', '', w)
#         if clean.isupper() and 2 <= len(clean) <= 6:
#             return clean
#     return None
# 
# def check_etf_status(security):
#     ticker = extract_ticker(security)
#     if not ticker:
#         return "NOT_ETF", None, None
#     try:
#         info = yf.Ticker(ticker).info
#         if info.get("quoteType") != "ETF":
#             return "NOT_ETF", ticker, None
#         aum = info.get("totalAssets")
#         if aum is None:
#             return "ETF_UNCERTAIN", ticker, None
#         if aum > 200_000_000:
#             return "ETF_EXEMPT", ticker, aum
#         return "ETF_CONTROL", ticker, aum
#     except:
#         return "NOT_ETF", ticker, None
# 
# # =========================
# # ✅ DATE RULE
# # =========================
# def is_late_trade(req_date, trade_date):
#     return trade_date > req_date + BDay(3)
# 
# # =========================
# # ✅ MAIN MATCHING
# # =========================
# results = []
# 
# for _, trade in trade_df.iterrows():
# 
#     staff = trade["Staff Name"]
#     trade_date = trade["Trade Date"]
#     trade_sec = trade["Security"]
#     trade_type = trade["Type"]
#     trade_qty = trade["Quantity"]
#     trade_isin = trade.get("ISIN")
# 
#     issues = []
# 
#     pad_filtered = pad_df[
#         (pad_df["Requestor"] == staff) &
#         (pad_df["Final Approval Date"] <= trade_date)
#     ].copy()
# 
#     if pad_filtered.empty:
#         results.append({**trade, "Issue": "No prior PAD approval"})
#         continue
# 
#     pad_filtered["Date Diff"] = (trade_date - pad_filtered["Final Approval Date"]).dt.days
#     pad_filtered = pad_filtered.sort_values("Date Diff").head(5)
# 
#     best_row = None
#     best_score = 0
# 
#     # ✅ ISIN MATCH
#     if pd.notna(trade_isin):
#         isin_match = pad_filtered[pad_filtered["ISIN Code"] == trade_isin]
#         if not isin_match.empty:
#             best_row = isin_match.iloc[0]
#             best_score = 100
# 
#     # ✅ FUZZY MATCH
#     if best_row is None:
#         for _, p in pad_filtered.iterrows():
#             pad_isin = p.get("ISIN Code")
#             if pd.notna(trade_isin) and pd.notna(pad_isin):
#                 continue
# 
#             score = fuzz.token_sort_ratio(
#                 enhance_name(trade_sec),
#                 enhance_name(p["Security"])
#             )
#             if score > best_score:
#                 best_score = score
#                 best_row = p
# 
#     if best_score < 70 or best_row is None:
#         results.append({
#             **trade,
#             "Matched Security": None,
#             "Match Score": best_score,
#             "Issue": "Unapproved Security"
#         })
#         continue
# 
#     req_date = best_row["Final Approval Date"]
#     req_type = best_row["Type"]
#     req_qty = best_row["Quantity"]
# 
#     if is_late_trade(req_date, trade_date):
#         issues.append("Late Trade")
# 
#     if str(trade_type).lower() != str(req_type).lower():
#         issues.append("Type Mismatch")
# 
#     if trade_qty > req_qty:
#         issues.append("Quantity Exceeded")
# 
#     results.append({
#         **trade,
#         "Matched Security": best_row["Security"],
#         "Match Score": best_score,
#         "Request Date": req_date,
#         "Approved Qty": req_qty,
#         "Issue": "OK" if not issues else "; ".join(issues)
#     })
# 
# df = pd.DataFrame(results)
# 
# # =========================
# # ✅ ✅ REVERSE CHECK (NEW CORE LOGIC)
# # =========================
# df["Comment"] = ""
# 
# for _, pad in pad_df.iterrows():
# 
#     staff = pad["Requestor"]
#     pad_sec = pad["Security"]
#     pad_date = pad["Final Approval Date"]
#     pad_isin = pad.get("ISIN Code")
# 
#     if pd.isna(pad_date):
#         continue
# 
#     trade_candidates = trade_df[
#         (trade_df["Staff Name"] == staff) &
#         (trade_df["Trade Date"] >= pad_date) &
#         (trade_df["Trade Date"] <= pad_date + BDay(3))
#     ]
# 
#     found_match = False
# 
#     # ✅ ISIN MATCH
#     if pd.notna(pad_isin):
#         if not trade_candidates[trade_candidates["ISIN"] == pad_isin].empty:
#             found_match = True
# 
#     # ✅ FUZZY FALLBACK
#     if not found_match:
#         for _, t in trade_candidates.iterrows():
#             score = fuzz.token_sort_ratio(
#                 enhance_name(pad_sec),
#                 enhance_name(t["Security"])
#             )
#             if score >= 70:
#                 found_match = True
#                 break
# 
#     # ✅ NO TRADE FOUND → ADD COMMENT
#     if not found_match:
# 
#         mask = (
#             (df["Staff Name"] == staff) &
#             (pd.to_datetime(df["Trade Date"], dayfirst=True, errors="coerce") >= pad_date)
#         )
# 
#         for idx in df[mask].index:
#             existing = df.at[idx, "Comment"]
#             new_text = f"Approved but not traded: {pad_sec}"
# 
#             if new_text not in existing:
#                 df.at[idx, "Comment"] = (
#                     existing + ("; " if existing else "") + new_text
#                 )
# 
# # =========================
# # ✅ FORMAT DATE
# # =========================
# df["Trade Date"] = pd.to_datetime(df["Trade Date"]).dt.strftime("%d-%m-%Y")
# 
# if "Request Date" in df.columns:
#     df["Request Date"] = pd.to_datetime(df["Request Date"], errors="coerce").dt.strftime("%d-%m-%Y")
# 
# # =========================
# # ✅ SAVE
# # =========================
# output_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\PAD sheets.xlsx"
# 
# with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
#     df.to_excel(writer, sheet_name="PAD Results", index=False)
# 
# # =========================
# # ✅ HIGHLIGHT
# # =========================
# wb = load_workbook(output_path)
# ws = wb["PAD Results"]
# 
# red = PatternFill(start_color="FF9999", fill_type="solid")
# 
# issue_col = [c.value for c in ws[1]].index("Issue") + 1
# 
# for r in range(2, ws.max_row + 1):
#     val = ws.cell(r, issue_col).value
#     if val and "Unapproved" in val:
#         for c in range(1, ws.max_column + 1):
#             ws.cell(r, c).fill = red
# 
# wb.save(output_path)
# 
# print("✅ FINAL — FULL BIDIRECTIONAL CHECK COMPLETE")

# # For submission records check

# import pandas as pd
# 
# 
# # ✅ Input file + sheet
# file_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\PAD sheets.xlsx"
# 
# # ✅ Output file (same file, different sheet)
# output_file = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\PAD sheets.xlsx"
# 
# 
# # ====================
# # ✅ LOAD DATA (READ from PAD Results)
# # ====================
# df = pd.read_excel(
#     file_path,
#     sheet_name="PAD Results"
# )
# 
# 
# # ✅ Convert date
# df["Trade Date"] = pd.to_datetime(
#     df["Trade Date"],
#     format="%d-%m-%Y",
#     errors="coerce"
# )
# 
# 
# # ✅ ✅ FIX 1: REMOVE INVALID DATES (avoid NaT column)
# df = df.dropna(subset=["Trade Date"])
# 
# # ✅ Extract Year-Month
# df["Year-Month"] = df["Trade Date"].dt.to_period("M").astype(str)
# 
# 
# # ====================
# # ✅ CREATE BASE TABLE (WITH ACCOUNT ID ✅)
# # ====================
# base = df[["Staff Name", "Account Name", "Account ID"]].drop_duplicates()
# 
# 
# # ====================
# # ✅ CREATE MONTH LIST (NO NaT ✅)
# # ====================
# all_months = sorted(df["Year-Month"].dropna().unique())
# 
# 
# # ====================
# # ✅ BUILD CONTROL MATRIX
# # ====================
# result = base.copy()
# 
# for month in all_months:
# 
#     temp = df[df["Year-Month"] == month][["Account ID"]].drop_duplicates()
# 
#     result[month] = result["Account ID"].isin(temp["Account ID"]).map({
#         True: "✅",
#         False: "❌"
#     })
# 
# 
# # ====================
# # ✅ SORT
# # ====================
# result = result.sort_values(by=["Staff Name", "Account Name"])
# 
# 
# # ====================
# # ✅ SAVE
# # ====================
# 
# with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
#     df.to_excel(writer, sheet_name="Statement Submission Tracker", index=False)
# 
# print("✅ DONE — Clean control file created")
# print("Saved to:", output_file)
# print(result)
# 

# In[5]:


import pandas as pd

# =========================
# ✅ FILE PATH
# =========================
output_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\PAD sheets.xlsx"
file_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\PAD sheets.xlsx"
registration_path = r"C:\Users\KateZhang\OneDrive - CIGP SA\Desktop\PAD Automation\2026 PAD statements submission tracker.xlsx"

# =========================
# ✅ LOAD DATA
# =========================
registration_df = pd.read_excel(registration_path, sheet_name="Statement Records")
statement_df = pd.read_excel(file_path, sheet_name="Extraction from Statement")

# =========================
# ✅ TARGET MONTH
# =========================
target_year = 2026
target_month = 6

# =========================
# ✅ CLEAN REGISTRATION (KEEP ALL ACCOUNTS ✅)
# =========================
registration = registration_df[[
    "Staff Names as per CIGP",
    "Staff Names as per Statements",
    "Broker",
    "A/C Number"
]].copy()   # ✅ no dropna here (important)

registration.columns = ["Staff", "StatementName", "Broker", "Account"]

registration["Staff"] = registration["Staff"].astype(str).str.strip()
registration["StatementName"] = registration["StatementName"].astype(str).str.strip().str.lower()
registration["Broker"] = registration["Broker"].astype(str).str.strip().str.lower()
registration["Account"] = registration["Account"].astype(str).str.strip()

# ✅ remove duplicates only
registration = registration.drop_duplicates()

# =========================
# ✅ CLEAN STATEMENT
# =========================
# =========================
# ✅ CLEAN REGISTRATION (FIX MERGED CELLS ✅)
# =========================
registration = registration_df[[
    "Staff Names as per CIGP",
    "Staff Names as per Statements",
    "Broker",
    "A/C Number"
]].copy()

registration.columns = ["Staff", "StatementName", "Broker", "Account"]

# ✅ 🔥 IMPORTANT: fill merged cells
registration["Staff"] = registration["Staff"].ffill()
registration["StatementName"] = registration["StatementName"].ffill()

# ✅ clean formatting
registration["Staff"] = registration["Staff"].astype(str).str.strip()
registration["StatementName"] = registration["StatementName"].astype(str).str.strip().str.lower()
registration["Broker"] = registration["Broker"].astype(str).str.strip().str.lower()
registration["Account"] = registration["Account"].astype(str).str.strip()

# ✅ remove rows without account (real filter)
registration = registration[registration["Account"] != ""]
registration = registration[registration["Account"].str.lower() != "nan"]

registration = registration.drop_duplicates()

# ✅ unique accounts only
statement = statement.drop_duplicates()

# =========================
# ✅ CREATE KEY (ACCOUNT-LEVEL ✅)
# =========================
registration["key"] = (
    registration["StatementName"] + "|" +
    registration["Broker"] + "|" +
    registration["Account"]
)

statement["key"] = (
    statement["StatementName"] + "|" +
    statement["Broker"] + "|" +
    statement["Account"]
)

# =========================
# ✅ FIND MISSING
# =========================
registration["Submitted"] = registration["key"].isin(statement["key"])

missing = registration[registration["Submitted"] == False].copy()

missing["Status"] = "Missing Statement"

# =========================
# ✅ OUTPUT
# =========================
result = missing[[
    "Staff",
    "StatementName",
    "Broker",
    "Account",
    "Status"
]].copy()

# ✅ fix warning
result.loc[:, "StatementName"] = result["StatementName"].str.title()
result.loc[:, "Broker"] = result["Broker"].str.title()

# =========================
# ✅ WRITE (DO NOT DELETE OTHER SHEETS ✅)
# =========================
with pd.ExcelWriter(
    output_path,
    engine="openpyxl",
    mode="a",                  # ✅ append
    if_sheet_exists="replace"  # ✅ only replace this sheet
) as writer:

    result.to_excel(writer, sheet_name="Missing Statements", index=False)

print("✅ Missing Statements sheet updated WITHOUT deleting other sheets")


# In[ ]:




